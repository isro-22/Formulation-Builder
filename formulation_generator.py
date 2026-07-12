from copy import copy
import csv
import io
import json
import re
from datetime import date as date_class, datetime
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
from typing import Any, Dict, Iterable, List, Optional

import openpyxl
import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parent
LOGO_PATH = PROJECT_ROOT / "assets" / "logo.png"
MATERIAL_DB_PATH = PROJECT_ROOT / "references" / "BOL-SAAT List.xlsx"
LOGO_MERGED_RANGE = "A1:B3"
EMPTY_SECTION_ROW_COUNT = 3
DEFAULT_EFFECTIVE_DATE = date_class(2025, 10, 10)
FORMULA_ROW_HEIGHT_PX = 15
FORMULA_ROW_HEIGHT_POINTS = 15.0
INPUT_SOURCE_FONT_COLOR = "7030A0"
PARENT_BOM_LEVEL1_FILL_COLOR = "E6D9F2"
PHASE_PERCENT_FORMAT = "0.00%"
USD_ACCOUNTING_FORMAT = '_([$$-409]* #,##0.00_);_([$$-409]* \\(#,##0.00\\);_([$$-409]* "-"??_);_(@_)'
DEFAULT_PREPARED_POSITION = "Flavourist"
DEFAULT_PREPARED_BY_NAME = "Lyla Isro"
DEFAULT_REVIEWED_BY = "Mochamad Setyawan"
DEFAULT_REVIEWED_POSITION = "Senior Manager - Flavourist"
DEFAULT_APPROVED_BY = "Andrew Yip"
DEFAULT_APPROVED_POSITION = "Head, Flavour PDI"


MAIN_PHASES = [
    "Casing Rajangan",
    "Casing Krosok",
    "Top Flavor",
]

INPUT_PHASE_METADATA_ROWS = MAIN_PHASES + [
    "Casing Pre-Mix",
    "Casing Pre-Mix 2",
    "Casing Pre-Mix 3",
    "Flavor Pre-Mix 1",
    "Flavor Pre-Mix 2",
    "Flavor Pre-Mix 3",
    "Flavor Pre-Mix 4",
    "Flavor Pre-Mix 5",
]

DEFAULT_PREMIX_PHASES = [
    "Casing Pre-Mix",
    "Flavor Pre-Mix 1",
    "Flavor Pre-Mix 2",
]

KNOWN_PHASES = MAIN_PHASES + DEFAULT_PREMIX_PHASES

DYNAMIC_FORMULA_PHASES = KNOWN_PHASES.copy()

BASE_SECTION_ROW_RANGES = {
    "Casing Rajangan": list(range(26, 31)),
    "Casing Krosok": list(range(39, 44)),
    "Top Flavor": list(range(52, 54)),
    "Casing Pre-Mix": list(range(77, 87)),
    "Flavor Pre-Mix 1": list(range(93, 128)),
    "Flavor Pre-Mix 2": list(range(134, 143)),
}

SECTION_ROW_RANGES = {phase: rows.copy() for phase, rows in BASE_SECTION_ROW_RANGES.items()}

BASE_PHASE_METADATA_POSITIONS = {
    "Casing Rajangan": {"nav_code": (21, 5), "description": (22, 5), "blend_ratio": (23, 5), "application": (24, 5)},
    "Casing Krosok": {"nav_code": (34, 5), "description": (35, 5), "blend_ratio": (36, 5), "application": (37, 5)},
    "Top Flavor": {"nav_code": (47, 5), "description": (48, 5), "blend_ratio": (49, 5), "application": (50, 5)},
    "Casing Pre-Mix": {"nav_code": (74, 5), "description": (75, 5), "blend_ratio": None, "application": None},
    "Flavor Pre-Mix 1": {"nav_code": (90, 5), "description": (91, 5), "blend_ratio": None, "application": None},
    "Flavor Pre-Mix 2": {"nav_code": (131, 5), "description": (132, 5), "blend_ratio": None, "application": None},
}

PHASE_METADATA_POSITIONS = {
    phase: positions.copy() for phase, positions in BASE_PHASE_METADATA_POSITIONS.items()
}

FORMULA_METADATA_POSITIONS = {
    "product_name": (5, 3),
    "prepared_by": (7, 3),
    "effective_date": (3, 17),
    "standard_control": (10, 3),
    "flavor_standard_reference": (10, 10),
    "tobacco_blend_code": (12, 3),
    "sensory_parameter": (12, 10),
    "formulation_code": (13, 3),
    "impact": (13, 10),
    "single_capsule": (14, 3),
    "flavor_aroma": (14, 10),
    "double_capsule_tobacco_end": (15, 3),
    "irritation": (15, 10),
    "double_capsule_mouth_end": (16, 3),
    "cooling": (16, 10),
    "product_weight_mg_stick": (18, 3),
    "clove_weight_mg_stick": (18, 6),
    "stick_per_mc": (18, 10),
}

INPUT_METADATA_POSITIONS = {
    "product_name": (2, 2),
    "formula_code": (3, 2),
    "product_weight_mg_stick": (4, 2),
    "clove_weight_mg_stick": (5, 2),
    "stick_per_mc": (6, 2),
    "prepared_by": (7, 2),
    "date": (8, 2),
    "approval_date": (9, 2),
    "standard_control": (3, 11),
    "flavor_standard_reference": (3, 14),
    "tobacco_blend_code": (5, 11),
    "sensory_parameter": (5, 14),
    "formulation_code": (6, 11),
    "impact": (6, 14),
    "single_capsule": (7, 11),
    "flavor_aroma": (7, 14),
    "double_capsule_tobacco_end": (8, 11),
    "irritation": (8, 14),
    "double_capsule_mouth_end": (9, 11),
    "cooling": (9, 14),
    "prepared_by_name": (11, 11),
    "prepared_position": (12, 11),
    "reviewed_by": (13, 11),
    "reviewed_position": (14, 11),
    "approved_by": (15, 11),
    "approved_position": (16, 11),
}

PHASE_METADATA_START_ROW = 4
PHASE_METADATA_ROW_COUNT = 14
MATERIAL_HEADER_ROW = 19
MATERIAL_START_ROW = 20
MATERIAL_COLUMNS = {
    "phase": 1,
    "item_code": 2,
    "item_name": 3,
    "dosage_input_mode": 4,
    "dosage_mg_stick": 5,
    "ratio_percent": 6,
    "addition_sequence": 7,
    "temperature": 8,
    "agitation_rate": 9,
    "mixing_duration": 10,
    "work_instruction_override": 11,
    "process_role": 12,
    "notes": 13,
}

NUMBERED_MATERIAL_COLUMNS = {
    "phase": 1,
    "no": 2,
    "item_code": 3,
    "item_name": 4,
    "dosage_input_mode": 5,
    "dosage_mg_stick": 6,
    "ratio_percent": 7,
    "addition_sequence": 8,
    "temperature": 9,
    "agitation_rate": 10,
    "mixing_duration": 11,
    "work_instruction_override": 12,
    "process_role": 13,
    "notes": 14,
}

INPUT_MATERIAL_HEADERS = {
    "phase": "Phase",
    "item_code": "Item Code",
    "item_name": "Item Name",
    "dosage_input_mode": "Input Mode",
    "dosage_mg_stick": "Dosage mg/stick",
    "ratio_percent": "Ratio %",
    "addition_sequence": "Addition Sequence",
    "temperature": "Temperature",
    "agitation_rate": "Agitation Rate",
    "mixing_duration": "Mixing Duration",
    "work_instruction_override": "Work Instruction Override",
    "process_role": "Process Role",
    "notes": "Notes",
}

FORMULA_MATERIAL_COLUMNS = {
    "no": 1,
    "item_code": 2,
    "item_name": 3,
    "physical_form": 5,
    "cas_number": 6,
    "ratio": 7,
    "dosage_mg_stick": 8,
    "material_price": 9,
    "formulation_price": 10,
    "dosage_kg_mc": 11,
    "density": 12,
    "addition_sequence": 13,
    "temperature": 14,
    "agitation_rate": 15,
    "mixing_duration": 16,
    "work_instruction": 17,
}

REQUIRED_FORMULATION_FIELDS = [
    "product_name",
    "formula_code",
    "prepared_by",
    "date",
    "product_weight_mg_stick",
    "clove_weight_mg_stick",
    "stick_per_mc",
]


@dataclass
class PhaseMetadata:
    phase: str
    nav_code: Optional[str] = None
    description: Optional[str] = None
    blend_ratio: Optional[float] = None
    application: Optional[float] = None


@dataclass
class MaterialInput:
    phase: str
    item_code: Optional[str]
    item_name: Optional[str]
    dosage_mg_stick: Optional[float]
    dosage_input_mode: str = "mg/stick"
    ratio_percent: Optional[float] = None
    application_percent: Optional[float] = None
    addition_sequence: Optional[int] = None
    temperature: Optional[str] = None
    agitation_rate: Optional[str] = None
    mixing_duration: Optional[str] = None
    work_instruction_override: Optional[str] = None
    process_role: Optional[str] = None
    notes: Optional[str] = None
    physical_form: Optional[str] = None
    cas_number: Optional[str] = None
    material_price: Optional[float] = None
    ratio: Optional[float] = None
    formulation_price: Optional[float] = None
    dosage_kg_mc: Optional[float] = None
    work_instruction: Optional[str] = None


@dataclass
class FormulationInput:
    product_name: str
    formula_code: str
    product_weight_mg_stick: float
    clove_weight_mg_stick: float
    stick_per_mc: int
    prepared_by: str
    prepared_by_name: str
    prepared_position: str
    date: Any
    approval_date: Any
    reviewed_by: str
    reviewed_position: str
    approved_by: str
    approved_position: str
    standard_control: str
    flavor_standard_reference: str
    tobacco_blend_code: str
    sensory_parameter: str
    formulation_code: str
    impact: str
    single_capsule: str
    flavor_aroma: str
    double_capsule_tobacco_end: str
    irritation: str
    double_capsule_mouth_end: str
    cooling: str
    phase_metadata: Dict[str, PhaseMetadata] = field(default_factory=dict)
    materials: List[MaterialInput] = field(default_factory=list)
    bypass_material_lookup: bool = False

    @property
    def effective_date(self) -> Any:
        return self.date

    @property
    def effective_approval_date(self) -> Any:
        return self.approval_date or self.effective_date


def normalize_phase(phase: str) -> str:
    if phase is None:
        raise ValueError("Phase name is required")
    phase_text = str(phase).strip()
    for known in KNOWN_PHASES:
        if phase_text.lower() == known.lower():
            return known
    lower_phase = phase_text.lower()
    for prefix in ("flavor pre-mix", "casing pre-mix"):
        if not lower_phase.startswith(prefix):
            continue
        suffix = phase_text[len(prefix):].strip()
        if not suffix:
            return "Casing Pre-Mix" if prefix == "casing pre-mix" else "Flavor Pre-Mix 1"
        try:
            number = int(suffix)
        except ValueError:
            break
        if number >= 1:
            if prefix == "casing pre-mix":
                return "Casing Pre-Mix" if number == 1 else f"Casing Pre-Mix {number}"
            return f"Flavor Pre-Mix {number}"
    return phase_text


def is_premix_phase(phase: str) -> bool:
    normalized = normalize_phase(phase)
    return normalized.startswith("Casing Pre-Mix") or normalized.startswith("Flavor Pre-Mix")


def phase_sort_key(phase: str) -> tuple:
    normalized = normalize_phase(phase)
    if normalized in MAIN_PHASES:
        return (0, MAIN_PHASES.index(normalized))
    if normalized.startswith("Casing Pre-Mix"):
        suffix = normalized.replace("Casing Pre-Mix", "").strip()
        number = int(suffix) if suffix.isdigit() else 1
        return (1, number)
    if normalized.startswith("Flavor Pre-Mix"):
        suffix = normalized.replace("Flavor Pre-Mix", "").strip()
        number = int(suffix) if suffix.isdigit() else 1
        return (2, number)
    return (9, normalized)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def is_missing_number(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def normalize_input_mode(value: Any) -> str:
    text = str(value or "mg/stick").strip().lower()
    if text in ("%", "percent", "percentage", "ratio/application %", "ratio %"):
        return "%"
    return "mg/stick"


def format_temperature_value(value: Any) -> Any:
    """Normalize numeric Celsius entries to a consistent Excel-friendly label."""
    if is_blank(value):
        return value
    text = str(value).strip()
    if text.lower() in {"ambient", "room temperature"}:
        return text
    match = re.fullmatch(r"([-+]?\d+(?:[.,]\d+)?)\s*(?:[°º]?\s*c(?:elsius)?)?", text, flags=re.IGNORECASE)
    if match:
        return f"{match.group(1)} °C"
    return re.sub(r"\s*[°º]?\s*c(?:elsius)?\b", " °C", text, flags=re.IGNORECASE)


def percent_to_fraction(value: Any) -> float:
    number = float(value)
    if abs(number) > 1:
        return number / 100
    return number


def round_dosage(value: float) -> float:
    return round(float(value), 5)


def effective_material_application_percent(formulation: FormulationInput, material: MaterialInput) -> Optional[float]:
    if material.application_percent not in (None, ""):
        return material.application_percent
    phase_metadata = formulation.phase_metadata.get(normalize_phase(material.phase))
    if phase_metadata and phase_metadata.application not in (None, ""):
        return phase_metadata.application
    return None


def effective_phase_blend_ratio(formulation: FormulationInput, phase: str) -> Optional[float]:
    phase_metadata = formulation.phase_metadata.get(normalize_phase(phase))
    if phase_metadata and phase_metadata.blend_ratio not in (None, ""):
        return phase_metadata.blend_ratio
    return None


def premix_parent_phases(phase: str) -> List[str]:
    normalized = normalize_phase(phase)
    if normalized.startswith("Casing Pre-Mix"):
        return ["Casing Rajangan", "Casing Krosok"]
    if normalized.startswith("Flavor Pre-Mix"):
        return ["Top Flavor"]
    return []


def material_matches_phase_metadata(material: "MaterialInput", metadata: "PhaseMetadata") -> bool:
    code = str(material.item_code or "").strip().lower()
    name = str(material.item_name or "").strip().lower()
    nav_code = str(metadata.nav_code or "").strip().lower()
    description = str(metadata.description or "").strip().lower()
    return bool(
        (nav_code and code == nav_code)
        or (description and name == description)
    )


def collect_premix_links(
    formulation: "FormulationInput",
    phase_groups: Dict[str, List["MaterialInput"]],
) -> tuple[Dict[str, List["MaterialInput"]], Dict[int, str]]:
    premix_to_parent_materials: Dict[str, List[MaterialInput]] = {}
    parent_material_to_premix: Dict[int, str] = {}

    for phase in phase_groups:
        if not is_premix_phase(phase):
            continue
        metadata = formulation.phase_metadata.get(normalize_phase(phase))
        if not metadata or (is_blank(metadata.nav_code) and is_blank(metadata.description)):
            continue

        matched_parents: List[MaterialInput] = []
        for parent_phase in premix_parent_phases(phase):
            for material in phase_groups.get(parent_phase, []):
                if material_matches_phase_metadata(material, metadata):
                    matched_parents.append(material)
                    parent_material_to_premix[id(material)] = phase

        if matched_parents:
            premix_to_parent_materials[phase] = matched_parents

    return premix_to_parent_materials, parent_material_to_premix


def phase_metadata_ref(row: int) -> str:
    return f"$E${row}"


def casing_blend_share_formula(phase: str) -> Optional[str]:
    positions = PHASE_METADATA_POSITIONS.get(phase)
    rajangan_positions = PHASE_METADATA_POSITIONS.get("Casing Rajangan")
    krosok_positions = PHASE_METADATA_POSITIONS.get("Casing Krosok")
    if not positions or not rajangan_positions or not krosok_positions:
        return None
    blend_position = positions.get("blend_ratio")
    rajangan_blend_position = rajangan_positions.get("blend_ratio")
    krosok_blend_position = krosok_positions.get("blend_ratio")
    if not blend_position or not rajangan_blend_position or not krosok_blend_position:
        return None
    blend_cell = phase_metadata_ref(blend_position[0])
    rajangan_blend_cell = phase_metadata_ref(rajangan_blend_position[0])
    krosok_blend_cell = phase_metadata_ref(krosok_blend_position[0])
    return f"({blend_cell}/({rajangan_blend_cell}+{krosok_blend_cell}))"


def percent_mode_dosage_formula(material: MaterialInput, row_index: int) -> Optional[str]:
    phase = normalize_phase(material.phase)
    positions = PHASE_METADATA_POSITIONS.get(phase, {})
    application_position = positions.get("application")
    if not application_position or material.ratio_percent in (None, ""):
        return None
    formula = f"=G{row_index}*($C$18-$F$18)*{phase_metadata_ref(application_position[0])}"
    if phase in ("Casing Rajangan", "Casing Krosok"):
        share_formula = casing_blend_share_formula(phase)
        if share_formula:
            formula = f"{formula}*{share_formula}"
    return formula


def reset_runtime_layout() -> None:
    SECTION_ROW_RANGES.clear()
    SECTION_ROW_RANGES.update({phase: rows.copy() for phase, rows in BASE_SECTION_ROW_RANGES.items()})
    PHASE_METADATA_POSITIONS.clear()
    PHASE_METADATA_POSITIONS.update({
        phase: positions.copy() for phase, positions in BASE_PHASE_METADATA_POSITIONS.items()
    })
    DYNAMIC_FORMULA_PHASES.clear()
    DYNAMIC_FORMULA_PHASES.extend(KNOWN_PHASES)


def shift_runtime_layout(insert_at: int, row_count: int, skip_phase: Optional[str] = None) -> None:
    for phase, rows in list(SECTION_ROW_RANGES.items()):
        if phase == skip_phase:
            continue
        SECTION_ROW_RANGES[phase] = [row + row_count if row >= insert_at else row for row in rows]

    for phase, positions in PHASE_METADATA_POSITIONS.items():
        shifted: Dict[str, Optional[tuple]] = {}
        for field_name, position in positions.items():
            if position is None:
                shifted[field_name] = None
                continue
            row, column = position
            shifted[field_name] = (row + row_count, column) if row >= insert_at else position
        PHASE_METADATA_POSITIONS[phase] = shifted


def desired_section_row_count(rows: List["MaterialInput"]) -> int:
    if rows:
        return len(rows)
    return EMPTY_SECTION_ROW_COUNT


def load_material_db(material_db_path: Path = MATERIAL_DB_PATH) -> pd.DataFrame:
    material_db_path = Path(material_db_path)
    if not material_db_path.exists():
        raise FileNotFoundError(f"Reference chemical database tidak ditemukan: {material_db_path}")
    df = pd.read_excel(material_db_path, sheet_name="BOL-SAAT List", engine="openpyxl")
    columns = {col: col for col in df.columns}
    normalized = {
        "item_code": columns.get("Item Code", "Item Code"),
        "item_name": columns.get("Item Name", "Item Name"),
        "price": columns.get("Price (USD)/KG", "Price (USD)/KG"),
        "cas_number": columns.get("CAS Number", "CAS Number"),
        "appearance": columns.get("Appearance", "Appearance"),
    }
    df = df.rename(columns={
        normalized["item_code"]: "item_code",
        normalized["item_name"]: "item_name",
        normalized["price"]: "price",
        normalized["cas_number"]: "cas_number",
        normalized["appearance"]: "appearance",
    })
    df = df.loc[:, ["item_code", "item_name", "price", "cas_number", "appearance"]]
    df["item_code"] = df["item_code"].astype(str).str.strip()
    df["item_name"] = df["item_name"].astype(str).str.strip()
    return df


def lookup_material_record(df: pd.DataFrame, item_code: Optional[str], item_name: Optional[str]) -> Optional[Dict[str, Any]]:
    if item_code:
        code = str(item_code).strip()
        row = df.loc[df["item_code"].str.lower() == code.lower()]
        if not row.empty:
            row = row.iloc[0]
            return row.to_dict()

    if item_name:
        name = str(item_name).strip().lower()
        row = df.loc[df["item_name"].str.lower() == name]
        if not row.empty:
            row = row.iloc[0]
            return row.to_dict()

    return None


def find_material_record(df: pd.DataFrame, item_code: Optional[str], item_name: Optional[str]) -> Dict[str, Any]:
    record = lookup_material_record(df, item_code, item_name)
    if record is not None:
        return record
    return {"item_code": item_code, "item_name": item_name, "price": None, "cas_number": None, "appearance": None}


def validate_formulation_input(formulation: FormulationInput, material_db: pd.DataFrame) -> List[str]:
    errors: List[str] = []

    for field_name in REQUIRED_FORMULATION_FIELDS:
        value = getattr(formulation, field_name)
        if is_blank(value):
            errors.append(f"Required field '{field_name}' wajib diisi.")

    try:
        if float(formulation.stick_per_mc) <= 0:
            errors.append("Field 'stick_per_mc' harus lebih dari 0.")
    except (TypeError, ValueError):
        errors.append("Field 'stick_per_mc' harus berupa angka.")

    if not formulation.materials:
        errors.append("Minimal satu material wajib diisi.")

    for index, material in enumerate(formulation.materials, start=1):
        row_label = f"Material row {index}"
        phase = normalize_phase(material.phase) if not is_blank(material.phase) else ""
        if not phase:
            errors.append(f"{row_label}: phase wajib diisi.")
        elif phase not in MAIN_PHASES and not is_premix_phase(phase):
            errors.append(
                f"{row_label}: phase '{phase}' tidak didukung. "
                "Gunakan Casing Rajangan, Casing Krosok, Top Flavor, atau pola Casing/Flavor Pre-Mix."
            )

        input_mode = normalize_input_mode(material.dosage_input_mode)
        if input_mode == "%":
            try:
                ratio = percent_to_fraction(material.ratio_percent)
                if ratio <= 0:
                    errors.append(f"{row_label}: ratio_percent harus lebih dari 0 untuk mode %.")
                if not is_premix_phase(phase):
                    application = percent_to_fraction(effective_material_application_percent(formulation, material))
                    if application <= 0:
                        errors.append(f"{row_label}: Application % harus lebih dari 0 untuk mode %.")
                if phase in ("Casing Rajangan", "Casing Krosok"):
                    rajangan_blend = percent_to_fraction(effective_phase_blend_ratio(formulation, "Casing Rajangan"))
                    krosok_blend = percent_to_fraction(effective_phase_blend_ratio(formulation, "Casing Krosok"))
                    if rajangan_blend <= 0 or krosok_blend <= 0 or rajangan_blend + krosok_blend <= 0:
                        errors.append(f"{row_label}: Blend Ratio Casing Rajangan dan Casing Krosok harus lebih dari 0 untuk mode %.")
            except (TypeError, ValueError):
                if is_premix_phase(phase):
                    errors.append(f"{row_label}: ratio_percent wajib angka untuk mode %.")
                else:
                    errors.append(f"{row_label}: ratio_percent, Application %, dan Blend Ratio casing wajib angka untuk mode %.")
        else:
            try:
                dosage = float(material.dosage_mg_stick)
                if dosage <= 0:
                    errors.append(f"{row_label}: dosage_mg_stick harus lebih dari 0.")
            except (TypeError, ValueError):
                errors.append(f"{row_label}: dosage_mg_stick harus berupa angka untuk mode mg/stick.")

        if is_blank(material.item_code) and is_blank(material.item_name):
            errors.append(f"{row_label}: item_code atau item_name wajib diisi.")
            continue

        record = lookup_material_record(material_db, material.item_code, material.item_name)
        if record is None:
            lookup_value = material.item_code or material.item_name
            if not formulation.bypass_material_lookup:
                errors.append(f"{row_label}: material lookup tidak ditemukan untuk '{lookup_value}'.")
            continue

        price = material.material_price if material.material_price is not None else record.get("price")
        if is_missing_number(price) and not formulation.bypass_material_lookup:
            lookup_value = material.item_code or material.item_name
            errors.append(f"{row_label}: price material kosong untuk '{lookup_value}'.")

    phases_with_materials = {normalize_phase(material.phase) for material in formulation.materials if not is_blank(material.phase)}
    for phase in MAIN_PHASES:
        if phase not in phases_with_materials:
            errors.append(f"Phase utama '{phase}' wajib memiliki minimal satu material.")

    return errors


def build_work_instruction(material: MaterialInput, phase_first: bool) -> str:
    if material.work_instruction_override:
        return material.work_instruction_override
    physical = material.physical_form
    if physical is None:
        physical = ""
    elif not isinstance(physical, str):
        physical = str(physical)
    physical = physical.strip().lower()
    if phase_first:
        if "solid" in physical:
            return "Magnetic stirrer. Add solids and mix until fully dissolved."
        return "Magnetic stirrer. Start mixing base materials till uniform."
    if material.process_role:
        return f"{material.process_role}. Continue mixing until homogeneous."
    if material.addition_sequence is not None:
        return f"Add material in sequence {material.addition_sequence}. Mix until uniform."
    return "Magnetic stirrer. Mix until homogeneous."


def compute_materials(formulation: FormulationInput, material_db: pd.DataFrame) -> None:
    phase_groups: Dict[str, List[MaterialInput]] = {}
    tobacco_weight_mg_stick = float(formulation.product_weight_mg_stick) - float(formulation.clove_weight_mg_stick)
    for material in formulation.materials:
        material.phase = normalize_phase(material.phase)
        material.dosage_input_mode = normalize_input_mode(material.dosage_input_mode)
        if material.dosage_input_mode == "%":
            ratio = percent_to_fraction(material.ratio_percent)
            if is_premix_phase(material.phase):
                material.dosage_mg_stick = None
            else:
                application = percent_to_fraction(effective_material_application_percent(formulation, material))
                if material.phase in ("Casing Rajangan", "Casing Krosok"):
                    rajangan_blend = percent_to_fraction(effective_phase_blend_ratio(formulation, "Casing Rajangan"))
                    krosok_blend = percent_to_fraction(effective_phase_blend_ratio(formulation, "Casing Krosok"))
                    phase_blend = percent_to_fraction(effective_phase_blend_ratio(formulation, material.phase))
                    application *= phase_blend / (rajangan_blend + krosok_blend)
                material.dosage_mg_stick = round_dosage(ratio * tobacco_weight_mg_stick * application)
        else:
            material.dosage_mg_stick = round_dosage(float(material.dosage_mg_stick or 0))
        phase_groups.setdefault(material.phase, []).append(material)

    premix_to_parent_materials, _ = collect_premix_links(formulation, phase_groups)
    for premix_phase, parent_materials in premix_to_parent_materials.items():
        premix_materials = phase_groups.get(premix_phase, [])
        if not premix_materials:
            continue

        parent_total_dosage = sum((material.dosage_mg_stick or 0.0) for material in parent_materials)
        for material in premix_materials:
            if material.dosage_input_mode == "%":
                material.dosage_mg_stick = round_dosage(percent_to_fraction(material.ratio_percent) * parent_total_dosage)

    for phase, materials in phase_groups.items():
        total_dosage = sum((m.dosage_mg_stick or 0) for m in materials)
        materials_sorted = sorted(materials, key=lambda m: (m.addition_sequence if m.addition_sequence is not None else 999, m.item_code or ""))
        for index, material in enumerate(materials_sorted, start=1):
            record = find_material_record(material_db, material.item_code, material.item_name)
            material.physical_form = material.physical_form or record.get("appearance")
            material.cas_number = material.cas_number or record.get("cas_number")
            material.material_price = material.material_price if material.material_price is not None else record.get("price")
            material.item_code = material.item_code or record.get("item_code")
            material.item_name = material.item_name or record.get("item_name")
            material.addition_sequence = material.addition_sequence or index
            material.ratio = 0.0 if total_dosage <= 0 else (material.dosage_mg_stick or 0) / total_dosage
            material.formulation_price = (material.ratio or 0.0) * (material.material_price or 0.0)
            material.dosage_kg_mc = (material.dosage_mg_stick or 0.0) * formulation.stick_per_mc / 1_000_000
            material.work_instruction = build_work_instruction(material, index == 1)

    for premix_phase, parent_materials in premix_to_parent_materials.items():
        premix_materials = phase_groups.get(premix_phase, [])
        premix_total_price = sum((material.formulation_price or 0.0) for material in premix_materials)
        for parent_material in parent_materials:
            parent_material.material_price = premix_total_price
            parent_material.formulation_price = (parent_material.ratio or 0.0) * premix_total_price

    formulation.materials = sorted(
        formulation.materials,
        key=lambda material: (
            phase_sort_key(material.phase),
            material.addition_sequence if material.addition_sequence is not None else 999999,
            str(material.item_code or material.item_name or ""),
        ),
    )


def writable_cell(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int):
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= column <= merged_range.max_col
            ):
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        return None
    return cell


def set_cell_value(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, column: int, value: Any) -> None:
    cell = writable_cell(ws, row, column)
    if cell is not None and not isinstance(cell, MergedCell):
        cell.value = value


def set_cell_reference_value(ws: openpyxl.worksheet.worksheet.Worksheet, cell_ref: str, value: Any) -> None:
    cell = ws[cell_ref]
    set_cell_value(ws, cell.row, cell.column, value)


def set_font_color(cell, color: str = "000000") -> None:
    if isinstance(cell, MergedCell):
        return
    font = copy(cell.font)
    font.color = color
    cell.font = font


def set_fill_color(cell, color: str) -> None:
    if isinstance(cell, MergedCell):
        return
    cell.fill = PatternFill("solid", fgColor=color)


def set_cell_locked(cell, locked: bool = True) -> None:
    if isinstance(cell, MergedCell):
        return
    protection = copy(cell.protection) if cell.protection is not None else Protection()
    protection.locked = locked
    cell.protection = protection


def clear_cells(ws: openpyxl.worksheet.worksheet.Worksheet, rows: Iterable[int], columns: Iterable[int]) -> None:
    for row in rows:
        for column in columns:
            set_cell_value(ws, row, column, None)


def copy_cell_format(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def copy_material_row_style(ws: openpyxl.worksheet.worksheet.Worksheet, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=column)
        target = ws.cell(row=target_row, column=column)
        if not isinstance(target, MergedCell):
            copy_cell_format(source, target)

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            ws.merge_cells(
                start_row=target_row,
                start_column=merged_range.min_col,
                end_row=target_row,
                end_column=merged_range.max_col,
            )


def copy_row_content_and_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = False
    for column in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=column)
        target = ws.cell(row=target_row, column=column)
        if isinstance(source, MergedCell) or isinstance(target, MergedCell):
            continue
        target.value = source.value
        copy_cell_format(source, target)


def copy_merged_ranges_for_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    source_start: int,
    source_end: int,
    target_start: int,
) -> None:
    row_offset = target_start - source_start
    existing_ranges = {str(merged_range) for merged_range in ws.merged_cells.ranges}
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row < source_start or merged_range.max_row > source_end:
            continue
        target_range = CellRange(
            min_col=merged_range.min_col,
            min_row=merged_range.min_row + row_offset,
            max_col=merged_range.max_col,
            max_row=merged_range.max_row + row_offset,
        )
        if str(target_range) not in existing_ranges:
            ws.merge_cells(str(target_range))
            existing_ranges.add(str(target_range))


def premix_source_phase(phase: str) -> str:
    return "Casing Pre-Mix" if normalize_phase(phase).startswith("Casing Pre-Mix") else "Flavor Pre-Mix 2"


def premix_insert_row(phase: str) -> int:
    normalized = normalize_phase(phase)
    if normalized.startswith("Casing Pre-Mix"):
        flavor_positions = [
            positions["nav_code"][0] - 1
            for phase_name, positions in PHASE_METADATA_POSITIONS.items()
            if phase_name.startswith("Flavor Pre-Mix") and positions.get("nav_code")
        ]
        if flavor_positions:
            return min(flavor_positions)
    premix_phases = [
        phase_name
        for phase_name in SECTION_ROW_RANGES
        if is_premix_phase(phase_name)
    ]
    return max(SECTION_ROW_RANGES[phase_name][-1] + 2 for phase_name in premix_phases)


def ensure_dynamic_premix_layouts(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    phase_rows: Dict[str, List[MaterialInput]],
) -> None:
    requested_phases = sorted(
        {normalize_phase(phase) for phase in phase_rows if is_premix_phase(phase)},
        key=phase_sort_key,
    )
    for phase in requested_phases:
        if phase in SECTION_ROW_RANGES:
            continue

        source_phase = premix_source_phase(phase)
        source_rows = SECTION_ROW_RANGES[source_phase]
        source_start = PHASE_METADATA_POSITIONS[source_phase]["nav_code"][0] - 1
        source_end = source_rows[-1] + 1
        block_height = source_end - source_start + 1
        insert_at = premix_insert_row(phase)

        ws.insert_rows(insert_at, amount=block_height)
        shift_runtime_layout(insert_at, block_height)

        shifted_source_start = source_start + block_height if source_start >= insert_at else source_start
        shifted_source_end = source_end + block_height if source_end >= insert_at else source_end
        for offset in range(block_height):
            copy_row_content_and_style(ws, shifted_source_start + offset, insert_at + offset)
        copy_merged_ranges_for_block(ws, shifted_source_start, shifted_source_end, insert_at)

        new_rows = [
            insert_at + (row - shifted_source_start)
            for row in range(shifted_source_start, shifted_source_end + 1)
            if shifted_source_start + 4 <= row <= shifted_source_end - 1
        ]
        SECTION_ROW_RANGES[phase] = new_rows
        PHASE_METADATA_POSITIONS[phase] = {
            "nav_code": (insert_at + 1, 5),
            "description": (insert_at + 2, 5),
            "blend_ratio": None,
            "application": None,
        }
        DYNAMIC_FORMULA_PHASES.append(phase)

        label_prefix = "Casing Pre-Mix Formulation" if phase.startswith("Casing Pre-Mix") else "Flavor Pre-Mix Formulation"
        set_cell_value(ws, insert_at, 1, label_prefix)
        set_cell_value(ws, insert_at + 1, 1, f"{label_prefix} NAV Item Code")
        set_cell_value(ws, insert_at + 2, 1, f"{label_prefix} NAV Item Description")


def unmerge_blocking_material_ranges(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    material_rows = {
        row
        for rows in SECTION_ROW_RANGES.values()
        for row in rows
    }
    for merged_range in list(ws.merged_cells.ranges):
        overlaps_material_rows = any(merged_range.min_row <= row <= merged_range.max_row for row in material_rows)
        overlaps_material_columns = merged_range.min_col <= 17 and merged_range.max_col >= 1
        if not overlaps_material_rows or not overlaps_material_columns:
            continue
        if merged_range.min_row == merged_range.max_row:
            if merged_range.min_col == 3 and merged_range.max_col == 4:
                continue

        source = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                target = ws.cell(row=row, column=column)
                copy_cell_format(source, target)


def enforce_material_name_merges(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    material_rows = {
        row
        for rows in SECTION_ROW_RANGES.values()
        for row in rows
    }

    for row in sorted(material_rows):
        for merged_range in list(ws.merged_cells.ranges):
            same_row_name_merge = (
                merged_range.min_row == row
                and merged_range.max_row == row
                and merged_range.min_col == 3
                and merged_range.max_col == 4
            )
            overlaps_name_cells = (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= 4
                and merged_range.max_col >= 3
            )
            if overlaps_name_cells and not same_row_name_merge:
                source = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                ws.unmerge_cells(str(merged_range))
                for unmerged_row in range(merged_range.min_row, merged_range.max_row + 1):
                    for column in range(merged_range.min_col, merged_range.max_col + 1):
                        copy_cell_format(source, ws.cell(row=unmerged_row, column=column))
        if f"C{row}:D{row}" in {str(merged_range) for merged_range in ws.merged_cells.ranges}:
            continue
        source = ws.cell(row=row, column=3)
        target = ws.cell(row=row, column=4)
        if isinstance(target, MergedCell):
            ws._cells.pop((row, 4), None)
            target = ws.cell(row=row, column=4)
        copy_cell_format(source, target)
        target.value = None
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)


def unmerge_ranges_overlapping_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    first_row: int,
    last_row: int,
) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row < first_row or merged_range.min_row > last_row:
            continue
        source = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                target = ws.cell(row=row, column=column)
                copy_cell_format(source, target)


def prepare_dynamic_formula_sections(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    phase_rows: Dict[str, List[MaterialInput]],
) -> None:
    for phase in DYNAMIC_FORMULA_PHASES:
        rows = phase_rows.get(phase, [])
        base_row_indices = SECTION_ROW_RANGES[phase]
        target_count = desired_section_row_count(rows)
        row_delta = target_count - len(base_row_indices)

        if row_delta > 0:
            insert_at = max(base_row_indices) + 1
            source_row = max(base_row_indices)
            ws.insert_rows(insert_at, amount=row_delta)
            for offset in range(row_delta):
                copy_material_row_style(ws, source_row, insert_at + offset)
            SECTION_ROW_RANGES[phase] = list(range(base_row_indices[0], base_row_indices[0] + target_count))
            shift_runtime_layout(insert_at, row_delta, skip_phase=phase)
            continue

        visible_rows = base_row_indices[:target_count]
        total_row = visible_rows[-1] + 1
        hidden_rows = [
            row
            for row in base_row_indices[target_count:]
            if row != total_row
        ]
        old_total_row = base_row_indices[-1] + 1
        if old_total_row != total_row:
            hidden_rows.append(old_total_row)
        SECTION_ROW_RANGES[phase] = visible_rows
        for row in hidden_rows:
            clear_cells(ws, [row], range(1, 18))
            ws.row_dimensions[row].hidden = True

    unmerge_blocking_material_ranges(ws)
    enforce_material_name_merges(ws)


def phase_header_labels(phase: str) -> Dict[str, str]:
    normalized = normalize_phase(phase)
    if normalized in ("Casing Rajangan", "Casing Krosok"):
        return {
            "title": normalized,
            "nav": "Casing NAV Item Code",
            "description": "Casing NAV Item Description",
            "blend": "Blend Ratio",
            "application": "Casing Application (%)",
            "physical": "Physical\nForm",
        }
    if normalized == "Top Flavor":
        return {
            "title": "Top Flavor",
            "nav": "Top Flavor NAV Item Code",
            "description": "Top Flavor Item Description",
            "blend": "Blend Ratio",
            "application": "TF Application (%)",
            "physical": "Physical State",
        }
    prefix = "Casing Pre-Mix Formulation" if normalized.startswith("Casing Pre-Mix") else "Flavor Pre-Mix Formulation"
    return {
        "title": prefix,
        "nav": f"{prefix} NAV Item Code",
        "description": f"{prefix} NAV Item Description",
        "physical": "Physical State",
    }


def rebuild_formula_section_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    table_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    material_headers = {
        1: "No",
        2: "Material NAV Item\nCode",
        3: "Material Name",
        5: None,
        6: "CAS Number",
        7: "Ratio\n(%)",
        8: "Dosage\n(mg/stick)",
        9: "Material\nPrice\n(USD / KG)",
        10: "Formulation\nPrice\n(USD / KG)",
        11: "Dosage\n(KG / MC)",
        12: "Density",
        13: "Addition\nSequence",
        14: "Temperature",
        15: "Agitation\nRate",
        16: "Mixing\nDuration",
        17: "Work Instruction",
    }

    def normal_cell(row: int, column: int):
        cell = ws.cell(row=row, column=column)
        if isinstance(cell, MergedCell):
            ws._cells.pop((row, column), None)
            cell = ws.cell(row=row, column=column)
        return cell

    for phase in DYNAMIC_FORMULA_PHASES:
        row_indices = SECTION_ROW_RANGES.get(phase)
        positions = PHASE_METADATA_POSITIONS.get(phase)
        if not row_indices or not positions or not positions.get("nav_code") or not positions.get("description"):
            continue

        first_material_row = row_indices[0]
        header_row = first_material_row - 1
        title_row = positions["nav_code"][0] - 1
        section_end = header_row
        unmerge_ranges_overlapping_rows(ws, title_row, section_end)
        clear_cells(ws, range(title_row, section_end + 1), range(1, 18))

        labels = phase_header_labels(phase)
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=17)
        title_cell = normal_cell(title_row, 1)
        title_cell.value = labels["title"]
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        metadata_rows = [
            (positions["nav_code"][0], labels["nav"]),
            (positions["description"][0], labels["description"]),
        ]
        if positions.get("blend_ratio"):
            metadata_rows.append((positions["blend_ratio"][0], labels["blend"]))
        if positions.get("application"):
            metadata_rows.append((positions["application"][0], labels["application"]))

        for row, label in metadata_rows:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=17)
            label_cell = normal_cell(row, 1)
            label_cell.value = label
            value_cell = normal_cell(row, 5)
            label_cell.font = Font(bold=True)
            label_cell.fill = label_fill
            set_font_color(value_cell)
            label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        material_headers[5] = labels["physical"]
        for column in range(1, 18):
            value = material_headers.get(column, "")
            cell = normal_cell(header_row, column)
            cell.value = value
            cell.font = Font(bold=True)
            cell.fill = table_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=4)

        for row in range(title_row, header_row + 1):
            ws.row_dimensions[row].hidden = False
            for column in range(1, 18):
                ws.cell(row=row, column=column).border = border


def build_formulation_code_formula() -> str:
    description_refs = []
    for phase in ("Casing Rajangan", "Casing Krosok", "Top Flavor"):
        description_position = PHASE_METADATA_POSITIONS[phase]["description"]
        description_refs.append(f"E{description_position[0]}")
    return f'=CONCATENATE({description_refs[0]},"/",{description_refs[1]},"/",{description_refs[2]})'


def build_input_formula_code_formula() -> str:
    return '=CONCATENATE(F4,"/",F5,"/",F6)'


def build_input_formulation_code_formula() -> str:
    return "=B3"


def format_approval_date(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d %B %Y")
    try:
        parsed = pd.to_datetime(value)
        if not pd.isna(parsed):
            return parsed.strftime("%d %B %Y")
    except (TypeError, ValueError):
        pass
    return str(value)


def apply_document_header_layout(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    formulation: FormulationInput,
) -> None:
    header_values = {
        "P1": "Document No",
        "Q1": "BOL-ID-BOLL-SPEC-003",
        "P2": "Revision No",
        "Q2": "00",
        "P3": "Effective Date",
        "Q3": formulation.effective_date,
        "A4": "Factory Lab to Fill Up",
        "A5": "Full FG Description",
        "A6": "Mixing Factory",
        "C6": "PT SAAT",
        "A7": "Lab Personnel Involved",
        "A9": "Flavor Development Reference",
        "A10": "Standard Control",
        "H10": "Flavor Standard Reference",
        "A11": "Product Specification",
        "A12": "Tobacco Blend Code",
        "H12": "Sensory Parameter",
        "A13": "Formulation Code",
        "H13": "Impact",
        "A14": "Single Capsule",
        "H14": "Flavor Aroma",
        "A15": "Double Capsule (Tobacco End)",
        "H15": "Irritation",
        "A16": "Double Capsule (Mouth End)",
        "H16": "Cooling",
    }
    for cell_ref, value in header_values.items():
        set_cell_reference_value(ws, cell_ref, value)

    ws["Q1"].font = Font(bold=True, color="FF0000")
    ws["Q2"].number_format = "@"
    for cell_ref in ("P1", "P2", "P3", "A4", "A5", "A6", "A7", "A9", "A10", "H10", "A11", "A12", "H12", "A13", "H13", "A14", "H14", "A15", "H15", "A16", "H16"):
        ws[cell_ref].font = copy(ws[cell_ref].font)
        ws[cell_ref].font = Font(bold=True)


def apply_approval_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    formulation: FormulationInput,
) -> None:
    approval_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Prepared By":
            approval_row = row
            break
    if approval_row is None:
        return

    cleanup_last_row = min(ws.max_row, approval_row + 9)
    unmerge_ranges_overlapping_rows(ws, approval_row, cleanup_last_row)
    clear_cells(ws, range(approval_row, cleanup_last_row + 1), range(1, 18))
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    title_fill = PatternFill("solid", fgColor="BFE3F0")
    white_fill = PatternFill(fill_type=None)
    empty_border = Border()
    for row in range(approval_row, cleanup_last_row + 1):
        ws.row_dimensions[row].height = 18
        for column in range(1, 18):
            cell = ws.cell(row=row, column=column)
            cell.fill = white_fill
            cell.border = empty_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[approval_row + 4].height = 58

    approval_date = format_approval_date(formulation.effective_approval_date)
    blocks = [
        (
            1,
            5,
            "Prepared By",
            f"Position : {formulation.prepared_position}\nName : {formulation.prepared_by_name}\nDate : {approval_date}",
        ),
        (
            6,
            12,
            "Reviewed By",
            f"Position : {formulation.reviewed_position}\nName : {formulation.reviewed_by}\nDate : {approval_date}",
        ),
        (
            13,
            17,
            "Approved By",
            f"Position : {formulation.approved_position}\nName : {formulation.approved_by}\nDate : {approval_date}",
        ),
    ]
    ws.row_dimensions[approval_row + 1].height = 48
    ws.row_dimensions[approval_row + 2].height = 54

    for start_column, end_column, title, details in blocks:
        ws.merge_cells(start_row=approval_row, start_column=start_column, end_row=approval_row, end_column=end_column)
        header_cell = ws.cell(row=approval_row, column=start_column)
        set_cell_value(ws, approval_row, start_column, title)
        header_cell.font = Font(bold=True)
        header_cell.fill = title_fill
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in range(start_column, end_column + 1):
            block_cell = ws.cell(row=approval_row, column=column)
            block_cell.border = border
            block_cell.fill = title_fill
            block_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells(start_row=approval_row + 1, start_column=start_column, end_row=approval_row + 1, end_column=end_column)
        detail_cell = ws.cell(row=approval_row + 1, column=start_column)
        set_cell_value(ws, approval_row + 1, start_column, details)
        detail_cell.fill = white_fill
        detail_cell.border = border
        detail_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for column in range(start_column, end_column + 1):
            block_cell = ws.cell(row=approval_row + 1, column=column)
            block_cell.border = border
            block_cell.fill = white_fill

        ws.merge_cells(start_row=approval_row + 2, start_column=start_column, end_row=approval_row + 2, end_column=end_column)
        signature_cell = ws.cell(row=approval_row + 2, column=start_column)
        set_cell_value(ws, approval_row + 2, start_column, "Signature:")
        signature_cell.fill = white_fill
        signature_cell.border = border
        signature_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for column in range(start_column, end_column + 1):
            block_cell = ws.cell(row=approval_row + 2, column=column)
            block_cell.border = border
            block_cell.fill = white_fill


def regenerate_total_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for phase, row_indices in SECTION_ROW_RANGES.items():
        first_row = row_indices[0]
        last_row = row_indices[-1]
        total_row = last_row + 1
        ws.row_dimensions[total_row].hidden = False
        unmerge_ranges_overlapping_rows(ws, total_row, total_row)
        clear_cells(ws, [total_row], range(1, 18))
        set_cell_value(ws, total_row, 1, "Total")
        set_cell_value(ws, total_row, 7, f"=SUM(G{first_row}:G{last_row})")
        set_cell_value(ws, total_row, 8, f"=SUM(H{first_row}:H{last_row})")
        set_cell_value(ws, total_row, 10, f"=SUM(J{first_row}:J{last_row})")
        set_cell_value(ws, total_row, 11, f"=SUM(K{first_row}:K{last_row})")


def apply_total_row_style(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    total_fill = PatternFill("solid", fgColor="D9D9D9")
    total_top = Side(style="thin", color="000000")
    total_bottom = Side(style="thin", color="000000")
    total_left = Side(style="thin", color="000000")
    total_right = Side(style="thin", color="000000")

    for row_indices in SECTION_ROW_RANGES.values():
        total_row = row_indices[-1] + 1
        for column in range(1, 18):
            cell = ws.cell(row=total_row, column=column)
            cell.fill = total_fill
            if column >= 7:
                cell.border = Border(
                    left=total_left,
                    right=total_right,
                    top=total_top,
                    bottom=total_bottom,
                )
                continue
            cell.border = Border(
                left=total_left if column == 1 else Side(style=None),
                right=Side(style=None),
                top=total_top,
                bottom=total_bottom,
            )


def apply_formula_number_formats(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    currency_format = USD_ACCOUNTING_FORMAT
    percent_format = "0.000%"
    dosage_format = "0.00000"
    white_fill = PatternFill(fill_type=None)

    for phase, row_indices in SECTION_ROW_RANGES.items():
        for row in row_indices:
            for column in range(1, 18):
                ws.cell(row=row, column=column).fill = white_fill
                ws.cell(row=row, column=column).alignment = Alignment(
                    horizontal="left" if column in (2, 3, 4, 5, 6, 17) else "center",
                    vertical="top",
                    wrap_text=True,
                )
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["ratio"]).number_format = percent_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).number_format = dosage_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["material_price"]).number_format = currency_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["formulation_price"]).number_format = currency_format
            ws.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["dosage_kg_mc"]).number_format = "0.00000"

        total_row = row_indices[-1] + 1
        ws.cell(row=total_row, column=7).number_format = percent_format
        ws.cell(row=total_row, column=8).number_format = dosage_format
        ws.cell(row=total_row, column=10).number_format = currency_format
        ws.cell(row=total_row, column=11).number_format = "0.00000"

    apply_total_row_style(ws)


def first_non_blank(values: Iterable[Any]) -> Any:
    for value in values:
        if not is_blank(value):
            return value
    return None


def merge_process_instruction_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    phase_rows: Dict[str, List[MaterialInput]],
) -> None:
    row_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    process_columns = [
        FORMULA_MATERIAL_COLUMNS["addition_sequence"],
        FORMULA_MATERIAL_COLUMNS["temperature"],
        FORMULA_MATERIAL_COLUMNS["agitation_rate"],
        FORMULA_MATERIAL_COLUMNS["mixing_duration"],
        FORMULA_MATERIAL_COLUMNS["work_instruction"],
    ]
    value_getters = {
        FORMULA_MATERIAL_COLUMNS["addition_sequence"]: lambda rows: rows[0].addition_sequence,
        FORMULA_MATERIAL_COLUMNS["temperature"]: lambda rows: first_non_blank(row.temperature for row in rows),
        FORMULA_MATERIAL_COLUMNS["agitation_rate"]: lambda rows: first_non_blank(row.agitation_rate for row in rows),
        FORMULA_MATERIAL_COLUMNS["mixing_duration"]: lambda rows: first_non_blank(row.mixing_duration for row in rows),
        FORMULA_MATERIAL_COLUMNS["work_instruction"]: lambda rows: (
            first_non_blank(row.work_instruction_override for row in rows)
            or first_non_blank(row.work_instruction for row in rows)
        ),
    }

    for phase, materials in phase_rows.items():
        row_indices = SECTION_ROW_RANGES.get(phase, [])
        if not row_indices or not materials:
            continue

        unmerge_ranges_overlapping_rows(ws, row_indices[0], row_indices[-1])

        group_start = 0
        while group_start < len(materials):
            sequence = materials[group_start].addition_sequence
            group_end = group_start + 1
            while group_end < len(materials) and materials[group_end].addition_sequence == sequence:
                group_end += 1

            start_row = row_indices[group_start]
            end_row = row_indices[group_end - 1]
            grouped_materials = materials[group_start:group_end]

            for column in process_columns:
                if is_premix_phase(phase):
                    for current_row, material in zip(row_indices[group_start:group_end], grouped_materials):
                        value = getattr(
                            material,
                            {
                                FORMULA_MATERIAL_COLUMNS["addition_sequence"]: "addition_sequence",
                                FORMULA_MATERIAL_COLUMNS["temperature"]: "temperature",
                                FORMULA_MATERIAL_COLUMNS["agitation_rate"]: "agitation_rate",
                                FORMULA_MATERIAL_COLUMNS["mixing_duration"]: "mixing_duration",
                                FORMULA_MATERIAL_COLUMNS["work_instruction"]: "work_instruction",
                            }[column],
                        )
                        if column == FORMULA_MATERIAL_COLUMNS["work_instruction"]:
                            value = material.work_instruction_override or material.work_instruction
                        set_cell_value(ws, current_row, column, value)
                        cell = writable_cell(ws, current_row, column)
                        if cell is None:
                            continue
                        cell.alignment = Alignment(
                            horizontal="center" if column == FORMULA_MATERIAL_COLUMNS["addition_sequence"] else "left",
                            vertical="center",
                            wrap_text=True,
                        )
                    for current_row in row_indices[group_start:group_end]:
                        for border_column in range(1, 18):
                            ws.cell(row=current_row, column=border_column).border = row_border
                    continue

                set_cell_value(ws, start_row, column, value_getters[column](grouped_materials))
                cell = writable_cell(ws, start_row, column)
                if cell is None:
                    continue
                cell.alignment = Alignment(
                    horizontal="center" if column == FORMULA_MATERIAL_COLUMNS["addition_sequence"] else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if column == FORMULA_MATERIAL_COLUMNS["work_instruction"]:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )
                if end_row > start_row:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=column,
                        end_row=end_row,
                        end_column=column,
                    )

            group_start = group_end


def update_top_formula_summary(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_row = None
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=2).value
        if isinstance(value, str) and "Casing & Top Flavor" in value:
            header_row = row
            break
    if header_row is None:
        return

    unmerge_ranges_overlapping_rows(ws, header_row, header_row + 5)
    clear_cells(ws, range(header_row, header_row + 6), range(1, 18))

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="D9D9D9")
    white_fill = PatternFill(fill_type=None)
    empty_border = Border()
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for row in range(header_row, header_row + 6):
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].height = 24
        for column in range(1, 18):
            cell = ws.cell(row=row, column=column)
            cell.fill = white_fill
            cell.border = empty_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        (1, "No"),
        (2, "Casing & Top Flavor\nNAV Item Code"),
        (3, "Application Amount\n(mg/stick)"),
        (4, "Blend Ratio"),
        (7, "Price\n(USD/KG)"),
        (8, "1,000 Sticks\nPrice\n(USD)"),
    ]
    for column, value in headers:
        set_cell_value(ws, header_row, column, value)
        cell = writable_cell(ws, header_row, column)
        if cell is None:
            continue
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "General"
    data_rows = []
    for index, phase in enumerate(("Casing Rajangan", "Casing Krosok", "Top Flavor"), start=1):
        row = header_row + index
        data_rows.append(row)
        positions = PHASE_METADATA_POSITIONS[phase]
        description = positions["description"]
        blend_ratio = positions["blend_ratio"]
        total_row = SECTION_ROW_RANGES[phase][-1] + 1

        set_cell_value(ws, row, 1, index)
        set_cell_value(ws, row, 2, f"=E{description[0]}")
        set_cell_value(ws, row, 3, f"=H{total_row}")
        set_cell_value(ws, row, 4, f"=E{blend_ratio[0]}")
        set_cell_value(ws, row, 7, f"=J{total_row}")
        set_cell_value(ws, row, 8, f"=G{row}*(C{row}/1000000)*1000")

        for column in (1, 2, 3, 4, 7, 8):
            cell = ws.cell(row=row, column=column)
            cell.fill = white_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if column == 2 else "center", vertical="center", wrap_text=True)
        ws.cell(row=row, column=3).number_format = "0.00000"
        ws.cell(row=row, column=4).number_format = "0%"
        ws.cell(row=row, column=7).number_format = USD_ACCOUNTING_FORMAT
        ws.cell(row=row, column=8).number_format = USD_ACCOUNTING_FORMAT

    total_row = header_row + 4
    clear_cells(ws, [total_row], range(1, 9))
    set_cell_value(ws, total_row, 7, "Total Price\n(USD/1,000 Sticks)")
    set_cell_value(ws, total_row, 8, f"=SUM(H{data_rows[0]}:H{data_rows[-1]})")
    for column in (7, 8):
        cell = ws.cell(row=total_row, column=column)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "General"
    ws.cell(row=total_row, column=8).number_format = USD_ACCOUNTING_FORMAT


def strip_xlsx_repair_risk_artifacts(workbook: openpyxl.Workbook) -> None:
    workbook._external_links = []
    for worksheet in workbook.worksheets:
        worksheet._images = []
        worksheet._charts = []


def apply_basic_sheet_style(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.sheet_view.showGridLines = True
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = copy(cell.alignment)


def find_summary_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=2).value
        if isinstance(value, str) and "Casing & Top Flavor" in value:
            return row
    return None


def find_approval_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Prepared By":
            return row
    return None


def apply_formula_row_height(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = FORMULA_ROW_HEIGHT_POINTS

    for row in range(1, 4):
        ws.row_dimensions[row].height = 20.0

    for row_indices in SECTION_ROW_RANGES.values():
        if row_indices:
            ws.row_dimensions[row_indices[0] - 1].height = 48.0

    summary_header_row = find_summary_header_row(ws)
    if summary_header_row:
        ws.row_dimensions[summary_header_row].height = 42
        for row in range(summary_header_row + 1, summary_header_row + 4):
            ws.row_dimensions[row].height = 28
        ws.row_dimensions[summary_header_row + 4].height = 30

    approval_row = find_approval_row(ws)
    if approval_row:
        ws.row_dimensions[approval_row].height = 18
        ws.row_dimensions[approval_row + 1].height = 48
        ws.row_dimensions[approval_row + 2].height = 54
        for row in range(approval_row + 3, approval_row + 5):
            ws.row_dimensions[row].height = FORMULA_ROW_HEIGHT_POINTS


def lock_formula_cells(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                set_cell_locked(cell, True)


def apply_parent_bom_level1_fill(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    phase_rows: Dict[str, List[MaterialInput]],
    parent_material_to_premix: Dict[int, str],
) -> None:
    for phase, rows in phase_rows.items():
        row_indices = SECTION_ROW_RANGES.get(phase, [])
        for row_index, material in zip(row_indices, rows):
            if not parent_material_to_premix.get(id(material)):
                continue
            for column in FORMULA_MATERIAL_COLUMNS.values():
                set_fill_color(ws.cell(row=row_index, column=column), PARENT_BOM_LEVEL1_FILL_COLOR)


def place_logo_in_merged_cell(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    logo_path: Path = LOGO_PATH,
    merged_range: str = LOGO_MERGED_RANGE,
) -> None:
    existing_ranges = {str(cell_range) for cell_range in ws.merged_cells.ranges}
    if merged_range not in existing_ranges:
        ws.merge_cells(merged_range)

    logo_cell = ws[merged_range.split(":")[0]]
    set_cell_value(ws, logo_cell.row, logo_cell.column, None)
    logo_cell.alignment = Alignment(horizontal="center", vertical="center")

    if not logo_path.exists():
        return

    try:
        logo = ExcelImage(str(logo_path))
        logo.width = 185
        logo.height = 48
        logo.anchor = TwoCellAnchor(
            editAs="twoCell",
            _from=AnchorMarker(col=0, row=0),
            to=AnchorMarker(col=2, row=3),
        )
        ws.add_image(logo)
    except Exception:
        pass


def setup_clean_formula_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, logo_path: Path = LOGO_PATH) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    table_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for column in range(1, 18):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 16
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["Q"].width = 38

    place_logo_in_merged_cell(ws, logo_path)
    ws.merge_cells("C1:O2")
    set_cell_reference_value(ws, "C1", "Laboratory Formulation Sheet\n& Lab Work Instruction")
    ws["C1"].font = Font(bold=True, size=16)
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    set_cell_reference_value(ws, "P1", "Document No")
    set_cell_reference_value(ws, "Q1", "BOL-ID-BOLL-SPEC-003")
    set_cell_reference_value(ws, "P2", "Revision No")
    set_cell_reference_value(ws, "Q2", "0")
    set_cell_reference_value(ws, "P3", "Effective Date")

    labels = {
        "A4": "Factory Lab to Fill Up",
        "A5": "Full FG Description",
        "A6": "Mixing Factory",
        "C6": "PT SAAT",
        "A7": "Lab Personnel Involved",
        "A9": "Flavor Development Reference",
        "A10": "Standard Control",
        "H10": "Flavor Standard Reference",
        "A12": "Product Specification",
        "A13": "Formulation Code",
        "H12": "Sensory Parameter",
        "H13": "Impact",
        "A14": "Single Capsule",
        "H14": "Flavor Aroma",
        "A15": "Double Capsule (Tobacco End)",
        "H15": "Irritation",
        "A16": "Double Capsule (Mouth End)",
        "H16": "Cooling",
        "A18": "Product Weight (mg/stick)",
        "E18": "Clove Weight (mg/stick)",
        "I18": "Stick per MC",
    }
    for cell_ref, value in labels.items():
        set_cell_reference_value(ws, cell_ref, value)
        ws[cell_ref].font = Font(bold=True)
        ws[cell_ref].fill = section_fill

    for cell in ws["P1:Q3"][0] + ws["P1:Q3"][1] + ws["P1:Q3"][2]:
        cell.border = border
    for row in range(4, 19):
        for column in range(1, 18):
            ws.cell(row=row, column=column).border = border

    material_headers = [
        "No",
        "Material NAV Item\nCode",
        "Material Name",
        "",
        "Physical\nForm",
        "CAS Number",
        "Ratio\n(%)",
        "Dosage\n(mg/stick)",
        "Material\nPrice\n(USD / KG)",
        "Formulation\nPrice\n(USD / KG)",
        "Dosage\n(KG / MC)",
        "Density",
        "Addition\nSequence",
        "Temperature",
        "Agitation\nRate",
        "Mixing\nDuration",
        "Work Instruction",
    ]

    phase_titles = {
        "Casing Rajangan": (20, "Casing", True),
        "Casing Krosok": (33, "Casing", True),
        "Top Flavor": (46, "Top Flavor", True),
        "Casing Pre-Mix": (73, "Casing Pre-Mix Formulation", False),
        "Flavor Pre-Mix 1": (89, "Flavor Pre-Mix Formulation", False),
        "Flavor Pre-Mix 2": (130, "Flavor Pre-Mix Formulation", False),
    }
    for phase, (title_row, label_prefix, has_ratio) in phase_titles.items():
        positions = PHASE_METADATA_POSITIONS[phase]
        set_cell_value(ws, title_row, 1, phase if phase in STATIC_PHASES else label_prefix)
        title_cell = writable_cell(ws, title_row, 1)
        if title_cell is not None:
            title_cell.font = Font(bold=True, color="FFFFFF")
            title_cell.fill = header_fill
        set_cell_value(ws, positions["nav_code"][0], 1, f"{label_prefix} NAV Item Code")
        set_cell_value(ws, positions["description"][0], 1, f"{label_prefix} NAV Item Description")
        if has_ratio:
            set_cell_value(ws, positions["blend_ratio"][0], 1, "Blend Ratio")
            set_cell_value(ws, positions["application"][0], 1, f"{label_prefix} Application (%)")

        header_row = SECTION_ROW_RANGES[phase][0] - 1
        for column, value in enumerate(material_headers, start=1):
            set_cell_value(ws, header_row, column, value)
            cell = writable_cell(ws, header_row, column)
            if cell is None:
                continue
            cell.font = Font(bold=True)
            cell.fill = table_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in SECTION_ROW_RANGES[phase] + [SECTION_ROW_RANGES[phase][-1] + 1]:
            for column in range(1, 18):
                cell = ws.cell(row=row, column=column)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22
    ws.freeze_panes = "A25"


def clone_template_sheet_layout(
    source_ws: openpyxl.worksheet.worksheet.Worksheet,
    target_ws: openpyxl.worksheet.worksheet.Worksheet,
    logo_path: Path = LOGO_PATH,
) -> None:
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)

    for column_letter, dimension in source_ws.column_dimensions.items():
        target_dimension = target_ws.column_dimensions[column_letter]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel

    for row_index, dimension in source_ws.row_dimensions.items():
        target_dimension = target_ws.row_dimensions[row_index]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel

    for row in source_ws.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target_ws.cell(
                row=source_cell.row,
                column=source_cell.column,
                value=source_cell.value,
            )
            copy_cell_format(source_cell, target_cell)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    if source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref

    place_logo_in_merged_cell(target_ws, logo_path)


def write_material_db_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    material_db_path: Path = MATERIAL_DB_PATH,
) -> None:
    material_db_path = Path(material_db_path)
    if not material_db_path.exists():
        raise FileNotFoundError(f"Reference chemical database tidak ditemukan: {material_db_path}")
    df = pd.read_excel(material_db_path, sheet_name="BOL-SAAT List", engine="openpyxl")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for column_index, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
    for row_index, row in enumerate(df.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.border = border
    for column in range(1, len(df.columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 22
    ws.freeze_panes = "A2"


def write_input_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, formulation: FormulationInput) -> None:
    set_cell_value(ws, 7, 1, "Lab Personnel Involved")
    for field_name, position in INPUT_METADATA_POSITIONS.items():
        value = getattr(formulation, field_name)
        set_cell_value(ws, position[0], position[1], value)

    set_cell_value(ws, INPUT_METADATA_POSITIONS["formula_code"][0], INPUT_METADATA_POSITIONS["formula_code"][1], build_input_formula_code_formula())
    set_cell_value(ws, INPUT_METADATA_POSITIONS["formulation_code"][0], INPUT_METADATA_POSITIONS["formulation_code"][1], build_input_formulation_code_formula())

    for field_name, column in MATERIAL_COLUMNS.items():
        set_cell_value(ws, MATERIAL_HEADER_ROW, column, INPUT_MATERIAL_HEADERS[field_name])

    for row_offset, phase in enumerate(INPUT_PHASE_METADATA_ROWS, start=PHASE_METADATA_START_ROW):
        metadata = formulation.phase_metadata.get(normalize_phase(phase))
        set_cell_value(ws, row_offset, 4, phase)
        if not metadata:
            continue
        set_cell_value(ws, row_offset, 5, metadata.nav_code)
        set_cell_value(ws, row_offset, 6, metadata.description)
        set_cell_value(ws, row_offset, 7, metadata.blend_ratio)
        set_cell_value(ws, row_offset, 8, metadata.application)
        blend_cell = writable_cell(ws, row_offset, 7)
        application_cell = writable_cell(ws, row_offset, 8)
        if blend_cell is not None:
            blend_cell.number_format = PHASE_PERCENT_FORMAT
        if application_cell is not None:
            application_cell.number_format = PHASE_PERCENT_FORMAT

    clear_until = max(ws.max_row, MATERIAL_START_ROW + len(formulation.materials) + 50)
    clear_cells(ws, range(MATERIAL_START_ROW, clear_until + 1), MATERIAL_COLUMNS.values())

    for offset, material in enumerate(formulation.materials, start=0):
        row_index = MATERIAL_START_ROW + offset
        for field_name, column in MATERIAL_COLUMNS.items():
            value = getattr(material, field_name, None)
            if value is None:
                continue
            if field_name == "dosage_mg_stick":
                set_cell_value(ws, row_index, column, float(value))
            elif field_name in ("ratio_percent", "application_percent"):
                set_cell_value(ws, row_index, column, float(value))
            elif field_name == "temperature":
                set_cell_value(ws, row_index, column, format_temperature_value(value))
            else:
                set_cell_value(ws, row_index, column, value)

    for row in range(MATERIAL_START_ROW, clear_until + 1):
        ratio_cell = writable_cell(ws, row, MATERIAL_COLUMNS["ratio_percent"])
        if ratio_cell is not None:
            ratio_cell.number_format = PHASE_PERCENT_FORMAT


def write_formula_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, formulation: FormulationInput) -> None:
    phase_rows: Dict[str, List[MaterialInput]] = {}
    for material in formulation.materials:
        phase_rows.setdefault(material.phase, []).append(material)

    ensure_dynamic_premix_layouts(ws, phase_rows)
    prepare_dynamic_formula_sections(ws, phase_rows)
    rebuild_formula_section_layout(ws)

    for field_name, position in FORMULA_METADATA_POSITIONS.items():
        value = getattr(formulation, field_name)
        set_cell_value(ws, position[0], position[1], value)

    apply_document_header_layout(ws, formulation)

    if formulation.phase_metadata:
        for phase, metadata in formulation.phase_metadata.items():
            phase_key = normalize_phase(phase)
            positions = PHASE_METADATA_POSITIONS.get(phase_key)
            if not positions:
                continue
            if metadata.nav_code and positions.get("nav_code"):
                set_cell_value(ws, positions["nav_code"][0], positions["nav_code"][1], metadata.nav_code)
            if metadata.description and positions.get("description"):
                set_cell_value(ws, positions["description"][0], positions["description"][1], metadata.description)
            if metadata.blend_ratio is not None and positions.get("blend_ratio"):
                set_cell_value(ws, positions["blend_ratio"][0], positions["blend_ratio"][1], metadata.blend_ratio)
                blend_cell = writable_cell(ws, positions["blend_ratio"][0], positions["blend_ratio"][1])
                if blend_cell is not None:
                    blend_cell.number_format = PHASE_PERCENT_FORMAT
            if metadata.application is not None and positions.get("application"):
                set_cell_value(ws, positions["application"][0], positions["application"][1], metadata.application)
                application_cell = writable_cell(ws, positions["application"][0], positions["application"][1])
                if application_cell is not None:
                    application_cell.number_format = PHASE_PERCENT_FORMAT

    set_cell_value(
        ws,
        FORMULA_METADATA_POSITIONS["formulation_code"][0],
        FORMULA_METADATA_POSITIONS["formulation_code"][1],
        build_formulation_code_formula(),
    )

    clear_cells(
        ws,
        [row for rows in SECTION_ROW_RANGES.values() for row in rows],
        FORMULA_MATERIAL_COLUMNS.values(),
    )

    premix_to_parent_materials, parent_material_to_premix = collect_premix_links(formulation, phase_rows)
    row_lookup: Dict[int, int] = {}
    for phase, rows in phase_rows.items():
        row_indices = SECTION_ROW_RANGES.get(phase, [])
        for row_index, material in zip(row_indices, rows):
            row_lookup[id(material)] = row_index

    premix_parent_dosage_formula: Dict[str, str] = {}
    for premix_phase, parent_materials in premix_to_parent_materials.items():
        refs = [f"$H${row_lookup[id(material)]}" for material in parent_materials if id(material) in row_lookup]
        if not refs:
            continue
        premix_parent_dosage_formula[premix_phase] = refs[0] if len(refs) == 1 else f"({' + '.join(refs)})".replace(" + ", "+")

    for phase, rows in phase_rows.items():
        if phase not in SECTION_ROW_RANGES:
            raise ValueError(f"Unsupported phase: {phase}. Gunakan salah satu phase: {', '.join(KNOWN_PHASES)}")
        row_indices = SECTION_ROW_RANGES[phase]
        for row_index, material in zip(row_indices, rows):
            total_row = row_indices[-1] + 1
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["no"], row_indices.index(row_index) + 1)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["item_code"], material.item_code)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["item_name"], material.item_name)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["physical_form"], material.physical_form)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["cas_number"], material.cas_number)
            if is_premix_phase(phase) and material.dosage_input_mode == "%":
                ratio_value = percent_to_fraction(material.ratio_percent)
            elif material.dosage_input_mode == "%":
                ratio_value = percent_to_fraction(material.ratio_percent)
            else:
                ratio_value = f"=IF($H${total_row}=0,0,H{row_index}/$H${total_row})"
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["ratio"], ratio_value)
            if is_premix_phase(phase) and phase in premix_parent_dosage_formula:
                dosage_formula = f"=G{row_index}*{premix_parent_dosage_formula[phase]}"
            else:
                dosage_formula = percent_mode_dosage_formula(material, row_index) if material.dosage_input_mode == "%" else None
            set_cell_value(
                ws,
                row_index,
                FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"],
                dosage_formula or material.dosage_mg_stick,
            )
            parent_premix_phase = parent_material_to_premix.get(id(material))
            if parent_premix_phase and parent_premix_phase in SECTION_ROW_RANGES:
                material_price_value = f"=J{SECTION_ROW_RANGES[parent_premix_phase][-1] + 1}"
            else:
                material_price_value = material.material_price
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["material_price"], material_price_value)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["formulation_price"], f"=G{row_index}*I{row_index}")
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["dosage_kg_mc"], f"=H{row_index}*$J$18/1000000")
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["density"], None)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["addition_sequence"], material.addition_sequence)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["temperature"], format_temperature_value(material.temperature))
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["agitation_rate"], material.agitation_rate)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["mixing_duration"], material.mixing_duration)
            set_cell_value(ws, row_index, FORMULA_MATERIAL_COLUMNS["work_instruction"], material.work_instruction)
            for column in FORMULA_MATERIAL_COLUMNS.values():
                set_font_color(ws.cell(row=row_index, column=column))
            source_column = (
                FORMULA_MATERIAL_COLUMNS["ratio"]
                if material.dosage_input_mode == "%"
                else FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]
            )
            set_font_color(ws.cell(row=row_index, column=source_column), INPUT_SOURCE_FONT_COLOR)

    regenerate_total_rows(ws)
    apply_formula_number_formats(ws)
    merge_process_instruction_columns(ws, phase_rows)
    enforce_material_name_merges(ws)
    apply_total_row_style(ws)
    update_top_formula_summary(ws)
    apply_approval_block(ws, formulation)
    apply_formula_row_height(ws)
    apply_parent_bom_level1_fill(ws, phase_rows, parent_material_to_premix)
    lock_formula_cells(ws)


def ensure_xlsx_output_path(output_path: Path) -> Path:
    if output_path.suffix.lower() != ".xlsx":
        return output_path.with_suffix(".xlsx")
    return output_path


def generate_formulation_workbook(
    formulation: FormulationInput,
    template_path: Path,
    output_path: Path,
    material_db_path: Path = MATERIAL_DB_PATH,
) -> Path:
    output_path = ensure_xlsx_output_path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    reset_runtime_layout()
    material_db = load_material_db(material_db_path)
    validation_errors = validate_formulation_input(formulation, material_db)
    if validation_errors:
        raise ValueError("Validasi formulasi gagal:\n- " + "\n- ".join(validation_errors))
    compute_materials(formulation, material_db)

    template_workbook = openpyxl.load_workbook(template_path, keep_vba=False, keep_links=False)
    if "Formula" not in template_workbook.sheetnames:
        raise ValueError("Template workbook harus memiliki sheet 'Formula'.")

    workbook = openpyxl.load_workbook(BytesIO(create_blank_input_workbook(template_path)))
    formula_sheet = workbook.create_sheet("Formula", 0)
    material_db_sheet = workbook.create_sheet("BOL-SAAT List", 1)
    clone_template_sheet_layout(template_workbook["Formula"], formula_sheet)
    write_material_db_sheet(material_db_sheet, material_db_path)

    write_input_sheet(workbook["INPUT"], formulation)
    write_formula_sheet(workbook["Formula"], formulation)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    workbook.save(output_path)
    return output_path


def create_blank_input_workbook(template_path: Path) -> bytes:
    workbook = openpyxl.Workbook()
    input_sheet = workbook.active
    input_sheet.title = "INPUT"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    table_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    locked_fill = PatternFill("solid", fgColor="000000")
    locked_font = Font(color="FFFFFF")

    def add_dropdown(range_ref: str, options: List[str], prompt: str) -> None:
        quoted_options = ",".join(options)
        validation = DataValidation(
            type="list",
            formula1=f'"{quoted_options}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid selection",
            error="Pilih nilai dari dropdown atau kosongkan jika tidak diperlukan.",
            promptTitle="Available options",
            prompt=prompt,
        )
        input_sheet.add_data_validation(validation)
        validation.add(range_ref)

    input_sheet.merge_cells("A1:N1")
    input_sheet["A1"] = "FORMULATION INPUT"
    input_sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    input_sheet["A1"].fill = header_fill
    input_sheet["A1"].alignment = Alignment(horizontal="center")
    input_sheet.row_dimensions[1].height = 20.0
    input_sheet.row_dimensions[3].height = 20.0

    input_sheet.merge_cells("D2:H2")
    input_sheet["D2"] = "PHASE / CASING / FLAVOR METADATA"
    input_sheet.merge_cells("J2:N2")
    input_sheet["J2"] = "FLAVOR DEVELOPMENT REFERENCE"
    input_sheet.merge_cells("J4:N4")
    input_sheet["J4"] = "PRODUCT SPECIFICATION"

    for cell_ref in ("D2", "J2", "J4"):
        cell = input_sheet[cell_ref]
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")

    metadata_labels = {
        (2, 1): "Product Name",
        (3, 1): "Formula Code",
        (4, 1): "Product Weight (mg/stick)",
        (5, 1): "Clove Weight (mg/stick)",
        (6, 1): "Stick per MC",
        (7, 1): "Lab Personnel Involved",
        (8, 1): "Effective Date",
        (9, 1): "Approval Date",
        (3, 10): "Standard Control",
        (3, 13): "Flavor Standard Reference",
        (5, 10): "Tobacco Blend Code",
        (5, 13): "Sensory Parameter",
        (6, 10): "Formulation Code",
        (6, 13): "Impact",
        (7, 10): "Single Capsule",
        (7, 13): "Flavor Aroma",
        (8, 10): "Double Capsule (Tobacco End)",
        (8, 13): "Irritation",
        (9, 10): "Double Capsule (Mouth End)",
        (9, 13): "Cooling",
    }
    for (row, column), label in metadata_labels.items():
        cell = input_sheet.cell(row=row, column=column, value=label)
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.border = border

    for row in range(2, 10):
        input_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    input_sheet.cell(row=8, column=2, value=DEFAULT_EFFECTIVE_DATE)
    input_sheet.cell(row=8, column=2).number_format = "yyyy-mm-dd"
    input_sheet.cell(row=9, column=2, value=DEFAULT_EFFECTIVE_DATE)
    input_sheet.cell(row=9, column=2).number_format = "yyyy-mm-dd"
    input_sheet.cell(row=3, column=2, value=build_input_formula_code_formula())
    for row in (3, 5, 6, 7, 8, 9):
        input_sheet.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
    input_sheet.cell(row=6, column=11, value=build_input_formulation_code_formula())

    input_sheet.merge_cells("J10:N10")
    input_sheet["J10"] = "APPROVAL METADATA"
    input_sheet["J10"].font = Font(bold=True)
    input_sheet["J10"].fill = section_fill
    input_sheet["J10"].alignment = Alignment(horizontal="center")
    approval_labels = {
        (11, 10): "Prepared By Name",
        (12, 10): "Prepared By Position",
        (13, 10): "Reviewed By Name",
        (14, 10): "Reviewed By Position",
        (15, 10): "Approved By Name",
        (16, 10): "Approved By Position",
    }
    approval_defaults = {
        11: DEFAULT_PREPARED_BY_NAME,
        12: DEFAULT_PREPARED_POSITION,
        13: DEFAULT_REVIEWED_BY,
        14: DEFAULT_REVIEWED_POSITION,
        15: DEFAULT_APPROVED_BY,
        16: DEFAULT_APPROVED_POSITION,
    }
    for row in range(11, 17):
        input_sheet.merge_cells(start_row=row, start_column=11, end_row=row, end_column=14)
    for (row, column), label in approval_labels.items():
        label_cell = input_sheet.cell(row=row, column=column, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = section_fill
        label_cell.border = border
        value_cell = input_sheet.cell(row=row, column=11, value=approval_defaults[row])
        value_cell.border = border

    phase_headers = ["Phase", "NAV Item Code", "NAV Item Description", "Blend Ratio", "Application %"]
    for offset, header in enumerate(phase_headers, start=4):
        cell = input_sheet.cell(row=3, column=offset, value=header)
        cell.font = Font(bold=True)
        cell.fill = table_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row_offset, phase in enumerate(INPUT_PHASE_METADATA_ROWS, start=PHASE_METADATA_START_ROW):
        input_sheet.cell(row=row_offset, column=4, value=phase)
        for column in range(4, 9):
            input_sheet.cell(row=row_offset, column=column).border = border
        input_sheet.cell(row=row_offset, column=7).number_format = PHASE_PERCENT_FORMAT
        input_sheet.cell(row=row_offset, column=8).number_format = PHASE_PERCENT_FORMAT
        if phase.startswith("Casing Pre-Mix") or phase.startswith("Flavor Pre-Mix"):
            for column in (7, 8):
                cell = input_sheet.cell(row=row_offset, column=column)
                cell.fill = locked_fill
                cell.font = locked_font

    for field_name, column in MATERIAL_COLUMNS.items():
        cell = input_sheet.cell(row=MATERIAL_HEADER_ROW, column=column, value=INPUT_MATERIAL_HEADERS[field_name])
        cell.font = Font(bold=True)
        cell.fill = table_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in range(MATERIAL_START_ROW, MATERIAL_START_ROW + 250):
        for column in MATERIAL_COLUMNS.values():
            input_sheet.cell(row=row, column=column).border = border
        input_sheet.cell(row=row, column=MATERIAL_COLUMNS["dosage_mg_stick"]).number_format = "0.00000"
        input_sheet.cell(row=row, column=MATERIAL_COLUMNS["ratio_percent"]).number_format = PHASE_PERCENT_FORMAT

    material_last_row = MATERIAL_START_ROW + 249
    phase_options = MAIN_PHASES + [
        "Casing Pre-Mix",
        "Casing Pre-Mix 2",
        "Casing Pre-Mix 3",
        "Flavor Pre-Mix 1",
        "Flavor Pre-Mix 2",
        "Flavor Pre-Mix 3",
        "Flavor Pre-Mix 4",
        "Flavor Pre-Mix 5",
    ]
    add_dropdown(f"A{MATERIAL_START_ROW}:A{material_last_row}", phase_options, "Pilih phase material.")
    add_dropdown(f"D{MATERIAL_START_ROW}:D{material_last_row}", ["mg/stick", "%"], "Pilih mode input dosage.")
    add_dropdown(f"H{MATERIAL_START_ROW}:H{material_last_row}", ["Ambient", "Room Temperature", "40 °C", "50 °C", "60 °C"], "Pilih temperature umum.")
    add_dropdown(f"I{MATERIAL_START_ROW}:I{material_last_row}", ["300 RPM", "500 RPM", "800 RPM", "1000 RPM"], "Pilih agitation rate umum.")
    add_dropdown(f"J{MATERIAL_START_ROW}:J{material_last_row}", ["5 min", "10 min", "15 min", "20 min", "30 min"], "Pilih mixing duration umum.")
    add_dropdown(
        f"L{MATERIAL_START_ROW}:L{material_last_row}",
        ["Carrier/Base", "Solid Dissolve", "Liquid Addition", "Paste/Extract Dispersion", "Final Homogenization"],
        "Pilih process role material.",
    )

    widths = {
        "A": 24,
        "B": 18,
        "C": 30,
        "D": 18,
        "E": 20,
        "F": 14,
        "G": 16,
        "H": 16,
        "I": 22,
        "J": 24,
        "K": 24,
        "L": 22,
        "M": 24,
        "N": 24,
    }
    for column, width in widths.items():
        input_sheet.column_dimensions[column].width = width

    editable_refs = [
        "B2", "B4", "B5", "B6", "B7", "B8", "B9",
        "K3", "N3", "K5", "N5", "N6", "K7", "N7", "K8", "N8", "K9", "N9",
        "K11", "K12", "K13", "K14", "K15", "K16",
    ]
    for cell_ref in editable_refs:
        set_cell_locked(input_sheet[cell_ref], False)
    for row in range(PHASE_METADATA_START_ROW, PHASE_METADATA_START_ROW + PHASE_METADATA_ROW_COUNT):
        set_cell_locked(input_sheet.cell(row=row, column=5), False)
        set_cell_locked(input_sheet.cell(row=row, column=6), False)
        phase_name = input_sheet.cell(row=row, column=4).value or ""
        if str(phase_name) in MAIN_PHASES:
            set_cell_locked(input_sheet.cell(row=row, column=7), False)
            set_cell_locked(input_sheet.cell(row=row, column=8), False)
    for row in range(MATERIAL_START_ROW, MATERIAL_START_ROW + 250):
        for column in MATERIAL_COLUMNS.values():
            set_cell_locked(input_sheet.cell(row=row, column=column), False)
    input_sheet.protection.sheet = True
    input_sheet.protection.enable()
    input_sheet.protection.formatCells = False
    input_sheet.protection.formatColumns = False
    input_sheet.protection.formatRows = False
    input_sheet.protection.insertColumns = False
    input_sheet.protection.insertRows = False
    input_sheet.protection.deleteColumns = False
    input_sheet.protection.deleteRows = False
    input_sheet.freeze_panes = f"A{MATERIAL_START_ROW}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_formulation_input_from_dict(data: Dict[str, Any], material_rows: Iterable[Dict[str, Any]]) -> FormulationInput:
    def clean_value(value: Any) -> Any:
        return None if is_blank(value) else value

    phase_metadata: Dict[str, PhaseMetadata] = {}
    for phase_name, values in (data.get("phase_metadata") or {}).items():
        if phase_name:
            phase_metadata[normalize_phase(phase_name)] = PhaseMetadata(
                phase=phase_name,
                nav_code=values.get("nav_code"),
                description=values.get("description"),
                blend_ratio=values.get("blend_ratio"),
                application=values.get("application"),
            )

    materials = []
    for row in material_rows:
        if not row.get("phase") or pd.isna(row.get("phase")):
            continue
        dosage = clean_value(row.get("dosage_mg_stick"))
        dosage_input_mode = clean_value(row.get("dosage_input_mode")) or "mg/stick"
        addition_sequence = clean_value(row.get("addition_sequence"))
        if normalize_input_mode(dosage_input_mode) != "%" and is_blank(dosage):
            continue
        materials.append(MaterialInput(
            phase=normalize_phase(row.get("phase")),
            item_code=clean_value(row.get("item_code")),
            item_name=clean_value(row.get("item_name")),
            dosage_mg_stick=float(dosage) if not is_blank(dosage) else None,
            dosage_input_mode=dosage_input_mode,
            ratio_percent=clean_value(row.get("ratio_percent")),
            application_percent=clean_value(row.get("application_percent")),
            addition_sequence=int(addition_sequence) if not is_blank(addition_sequence) else None,
            temperature=format_temperature_value(clean_value(row.get("temperature"))),
            agitation_rate=clean_value(row.get("agitation_rate")),
            mixing_duration=clean_value(row.get("mixing_duration")),
            work_instruction_override=clean_value(row.get("work_instruction_override")),
            process_role=clean_value(row.get("process_role")),
            notes=clean_value(row.get("notes")),
        ))

    return FormulationInput(
        product_name=data.get("product_name", ""),
        formula_code=data.get("formula_code", ""),
        product_weight_mg_stick=float(data.get("product_weight_mg_stick", 0)),
        clove_weight_mg_stick=float(data.get("clove_weight_mg_stick", 0)),
        stick_per_mc=int(data.get("stick_per_mc", 10000)),
        prepared_by=data.get("prepared_by", ""),
        prepared_by_name=data.get("prepared_by_name") or DEFAULT_PREPARED_BY_NAME,
        prepared_position=data.get("prepared_position") or DEFAULT_PREPARED_POSITION,
        date=data.get("date") or DEFAULT_EFFECTIVE_DATE.isoformat(),
        approval_date=data.get("approval_date") or data.get("date") or DEFAULT_EFFECTIVE_DATE.isoformat(),
        reviewed_by=data.get("reviewed_by") or DEFAULT_REVIEWED_BY,
        reviewed_position=data.get("reviewed_position") or DEFAULT_REVIEWED_POSITION,
        approved_by=data.get("approved_by") or DEFAULT_APPROVED_BY,
        approved_position=data.get("approved_position") or DEFAULT_APPROVED_POSITION,
        standard_control=data.get("standard_control", ""),
        flavor_standard_reference=data.get("flavor_standard_reference", ""),
        tobacco_blend_code=data.get("tobacco_blend_code", ""),
        sensory_parameter=data.get("sensory_parameter", ""),
        formulation_code=data.get("formulation_code", ""),
        impact=data.get("impact", ""),
        single_capsule=data.get("single_capsule", ""),
        flavor_aroma=data.get("flavor_aroma", ""),
        double_capsule_tobacco_end=data.get("double_capsule_tobacco_end", ""),
        irritation=data.get("irritation", ""),
        double_capsule_mouth_end=data.get("double_capsule_mouth_end", ""),
        cooling=data.get("cooling", ""),
        phase_metadata=phase_metadata,
        materials=materials,
        bypass_material_lookup=bool(data.get("bypass_material_lookup", False)),
    )


def _serialize_model_value(value: Any) -> Any:
    if isinstance(value, (datetime, date_class)):
        return value.isoformat()
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    return value


def normalize_identifier(value: Any) -> str:
    text = str(value or "").strip()
    normalized = re.sub(r"[^A-Za-z0-9]+", "-", text).strip("-").upper()
    return normalized or "UNSPECIFIED"


def product_id_for(formulation: FormulationInput) -> str:
    return normalize_identifier(formulation.formula_code or formulation.product_name)


def phase_id_for(product_id: str, phase: Any) -> str:
    return f"{product_id}::{normalize_identifier(phase)}"


def export_formulation_model(formulation: FormulationInput) -> Dict[str, Any]:
    product_id = product_id_for(formulation)
    product = {
        "product_id": product_id,
        "product_name": formulation.product_name,
        "formula_code": formulation.formula_code,
        "product_weight_mg_stick": formulation.product_weight_mg_stick,
        "clove_weight_mg_stick": formulation.clove_weight_mg_stick,
        "stick_per_mc": formulation.stick_per_mc,
        "prepared_by": formulation.prepared_by,
        "date": _serialize_model_value(formulation.date),
        "approval_date": _serialize_model_value(formulation.effective_approval_date),
        "standard_control": formulation.standard_control,
        "flavor_standard_reference": formulation.flavor_standard_reference,
        "tobacco_blend_code": formulation.tobacco_blend_code,
        "sensory_parameter": formulation.sensory_parameter,
        "formulation_code": formulation.formulation_code,
        "impact": formulation.impact,
        "single_capsule": formulation.single_capsule,
        "flavor_aroma": formulation.flavor_aroma,
        "double_capsule_tobacco_end": formulation.double_capsule_tobacco_end,
        "irritation": formulation.irritation,
        "double_capsule_mouth_end": formulation.double_capsule_mouth_end,
        "cooling": formulation.cooling,
        "prepared_position": formulation.prepared_position,
        "reviewed_by": formulation.reviewed_by,
        "reviewed_position": formulation.reviewed_position,
        "approved_by": formulation.approved_by,
        "approved_position": formulation.approved_position,
        "lab_personnel_involved": formulation.prepared_by,
        "prepared_by_name": formulation.prepared_by_name,
        "bypass_material_lookup": formulation.bypass_material_lookup,
    }

    phase_metadata = [
        {
            "phase_id": phase_id_for(product_id, metadata.phase),
            "product_id": product_id,
            "phase": metadata.phase,
            "nav_code": metadata.nav_code,
            "description": metadata.description,
            "blend_ratio": metadata.blend_ratio,
            "application": metadata.application,
        }
        for metadata in sorted(formulation.phase_metadata.values(), key=lambda item: phase_sort_key(item.phase))
    ]

    materials = []
    sorted_materials = sorted(
        formulation.materials,
        key=lambda item: (
            phase_sort_key(item.phase),
            item.addition_sequence if item.addition_sequence is not None else 999999,
            item.item_code or "",
        ),
    )
    for material_index, material in enumerate(sorted_materials, start=1):
        phase_id = phase_id_for(product_id, material.phase)
        materials.append({
            "material_id": f"{phase_id}::{material_index:04d}",
            "product_id": product_id,
            "phase_id": phase_id,
            "phase": material.phase,
            "item_code": material.item_code,
            "item_name": material.item_name,
            "dosage_input_mode": material.dosage_input_mode,
            "dosage_mg_stick": material.dosage_mg_stick,
            "ratio_percent": material.ratio_percent,
            "addition_sequence": material.addition_sequence,
            "temperature": material.temperature,
            "agitation_rate": material.agitation_rate,
            "mixing_duration": material.mixing_duration,
            "work_instruction_override": material.work_instruction_override,
            "process_role": material.process_role,
            "notes": material.notes,
        })

    return {
        "schema_version": "1.0",
        "model_type": "formulation_product",
        "product": product,
        "phase_metadata": phase_metadata,
        "materials": materials,
    }


def export_formulation_model_json(formulation: FormulationInput, *, indent: int = 2) -> bytes:
    model = export_formulation_model(formulation)
    return json.dumps(model, ensure_ascii=False, indent=indent, default=str).encode("utf-8")


def export_formulation_model_xlsx(formulation: FormulationInput) -> bytes:
    model = export_formulation_model(formulation)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame([model["product"]]).to_excel(writer, sheet_name="product", index=False)
        pd.DataFrame(model["phase_metadata"]).to_excel(writer, sheet_name="phase_metadata", index=False)
        pd.DataFrame(model["materials"]).to_excel(writer, sheet_name="materials", index=False)

        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            for column_cells in worksheet.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)

    buffer.seek(0)
    return buffer.getvalue()


def _dicts_to_csv_bytes(rows: List[Dict[str, Any]], fieldnames: List[str]) -> bytes:
    text_io = io.StringIO()
    writer = csv.DictWriter(text_io, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({field: _serialize_model_value(row.get(field)) for field in fieldnames})
    return text_io.getvalue().encode("utf-8")


def export_formulation_model_csv_bundle(formulation: FormulationInput) -> bytes:
    model = export_formulation_model(formulation)
    product_rows = [model["product"]]
    phase_rows = model["phase_metadata"]
    material_rows = model["materials"]

    csv_files = {
        "product.csv": _dicts_to_csv_bytes(product_rows, list(model["product"].keys())),
        "phase_metadata.csv": _dicts_to_csv_bytes(
            phase_rows,
            ["phase", "nav_code", "description", "blend_ratio", "application"],
        ),
        "materials.csv": _dicts_to_csv_bytes(
            material_rows,
            [
                "phase",
                "item_code",
                "item_name",
                "dosage_input_mode",
                "dosage_mg_stick",
                "ratio_percent",
                "addition_sequence",
                "temperature",
                "agitation_rate",
                "mixing_duration",
                "work_instruction_override",
                "process_role",
                "notes",
            ],
        ),
    }

    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        for filename, content in csv_files.items():
            archive.writestr(filename, content)
    buffer.seek(0)
    return buffer.getvalue()
