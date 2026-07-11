"""
template_scraper.py
====================
Aplikasi Streamlit untuk scraping data formulasi dari file Excel
template original (contoh: template_1.xlsx) dan menyimpannya ke
database JSON dan XLSX.

Konsep:
- Baca file Excel template yang sudah ada (format asli / legacy)
- Deteksi otomatis section: header, phase metadata, material table, premix
- Ekstrak semua data yang bisa dibaca (abaikan sel #REF! / formula rusak)
- Simpan ke JSON terstruktur dan XLSX normalisasi (3 sheet: product, phase_metadata, materials)
- Tampilkan preview hasil scraping di Streamlit
"""

import io
import json
import re
from datetime import date as date_class, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# Konstanta
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_ROOT / "outputs" / "scraped"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SCHEMA_VERSION = "1.0"

# Kolom yang dibaca dari baris material (sesuai layout template_1.xlsx)
MATERIAL_COL_NO = 1
MATERIAL_COL_ITEM_CODE = 2
MATERIAL_COL_ITEM_NAME = 3
MATERIAL_COL_PHYSICAL_FORM = 5
MATERIAL_COL_CAS = 6
MATERIAL_COL_RATIO = 7
MATERIAL_COL_DOSAGE = 8
MATERIAL_COL_MAT_PRICE = 9
MATERIAL_COL_FORM_PRICE = 10
MATERIAL_COL_DOSAGE_KG = 11
MATERIAL_COL_DENSITY = 12
MATERIAL_COL_SEQ = 13
MATERIAL_COL_TEMP = 14
MATERIAL_COL_AGIT = 15
MATERIAL_COL_MIX_DUR = 16
MATERIAL_COL_WI = 17


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _clean(value: Any) -> Any:
    """Bersihkan nilai: ubah #REF! / string error menjadi None."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", ""}:
            return None
        return stripped
    if isinstance(value, float) and (value != value):  # NaN
        return None
    return value


def _clean_ratio(value: Any) -> Optional[float]:
    """Ratio di template disimpan sebagai desimal (0.38 = 38%). Kembalikan sebagai float."""
    cleaned = _clean(value)
    if cleaned is None:
        return None
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def compute_dosage_mg_stick(
    ratio_raw: Optional[float],
    product_weight_mg_stick: Optional[float],
    phase_app_pct: Optional[float] = None,
    phase_blend_ratio: Optional[float] = None,
) -> Optional[float]:
    """
    Hitung dosage mg/stick dari ratio internal fase.

    Formula:
        dosage = ratio_internal × application_mg_stick

    Di mana:
        application_mg_stick = product_weight × (application_pct / 100)
                                jika application_pct tersedia (Casing / Top Flavor)
        application_mg_stick = product_weight × (blend_ratio / 100)
                                fallback jika hanya blend_ratio tersedia

    Catatan: Untuk Premix, dosage dihitung dari ratio premix × total TF dosage
    (tidak diimplementasikan di sini — cukup simpan ratio_pct saja).
    """
    if ratio_raw is None or product_weight_mg_stick is None:
        return None
    # application_pct sudah dalam % (misal 3.078), konversi ke desimal
    if phase_app_pct is not None:
        app_fraction = phase_app_pct / 100.0
    elif phase_blend_ratio is not None:
        app_fraction = phase_blend_ratio / 100.0
    else:
        return None
    dosage = ratio_raw * (product_weight_mg_stick * app_fraction)
    return round(dosage, 6)


def _is_header_row(row_values: List[Any]) -> bool:
    """Deteksi baris header material table (kolom 1 = 'No', kolom 2 = 'Material NAV...')."""
    col1 = _clean(row_values[0])
    col2 = _clean(row_values[1])
    return (
        isinstance(col1, str) and col1.strip().lower() == "no"
        and isinstance(col2, str) and "material nav" in col2.strip().lower()
    )


def _is_total_row(row_values: List[Any]) -> bool:
    """Deteksi baris Total."""
    col1 = _clean(row_values[0])
    return isinstance(col1, str) and col1.strip().lower() == "total"


def _is_material_row(row_values: List[Any]) -> bool:
    """Deteksi baris material: kolom 1 adalah angka (nomor urut)."""
    col1 = _clean(row_values[0])
    if col1 is None:
        return False
    try:
        num = int(float(str(col1)))
        return num > 0
    except (ValueError, TypeError):
        return False


def _parse_phase_metadata_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    phase_label: str,
) -> Tuple[Dict[str, Any], int]:
    """
    Baca 5-6 baris metadata phase (NAV code, description, blend ratio, application).
    Return (metadata_dict, next_row_index).
    """
    metadata: Dict[str, Any] = {
        "phase": phase_label,
        "nav_code": None,
        "description": None,
        "blend_ratio": None,
        "application": None,
    }

    search_rows = range(start_row, start_row + 10)
    for r in search_rows:
        col1 = _clean(ws.cell(row=r, column=1).value)
        col5 = _clean(ws.cell(row=r, column=5).value)
        if col1 is None:
            continue
        label = str(col1).strip().lower()

        # NAV Item Code — berbagai variasi label:
        # "Casing NAV Item Code", "Top Flavor NAV Item Code",
        # "Flavor Pre-Mix Formulation NAV Item Code"
        if "nav item code" in label or "item code" in label:
            if metadata["nav_code"] is None:  # ambil yang pertama
                metadata["nav_code"] = col5

        # Description — variasi: "Casing NAV Item Description",
        # "Top Flavor Item Description", "Flavor Pre-Mix ... NAV Item Description"
        elif "item description" in label:
            if metadata["description"] is None:
                metadata["description"] = col5

        # Blend Ratio (sama di semua fase)
        elif "blend ratio" in label:
            v = _clean_ratio(ws.cell(row=r, column=5).value)
            metadata["blend_ratio"] = round(v * 100, 6) if v is not None else None

        # Application — variasi: "Casing Application (%)", "TF Application (%)"
        elif "application" in label:
            v = _clean_ratio(ws.cell(row=r, column=5).value)
            metadata["application"] = round(v * 100, 6) if v is not None else None

    # Temukan baris header material setelah blok metadata
    # Cari lebih lebar (s/d +12) untuk mengakomodasi variasi panjang blok header
    header_row = start_row + 6
    for r in range(start_row, start_row + 12):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if _is_header_row(vals):
            header_row = r
            break

    return metadata, header_row + 1  # material data dimulai baris setelah header


def _read_material_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    phase_name: str,
    max_rows: int = 200,
    product_weight_mg_stick: Optional[float] = None,
    phase_app_pct: Optional[float] = None,
    phase_blend_ratio: Optional[float] = None,
) -> Tuple[List[Dict[str, Any]], int]:
    """
    Baca baris material dari start_row hingga bertemu baris Total.
    Jika dosage_mg_stick adalah #REF! dan product_weight + application tersedia,
    hitung otomatis dari ratio × product_weight × application_pct.
    Return (list_of_materials, row_after_total).
    """
    materials: List[Dict[str, Any]] = []
    current_temp = None
    current_agit = None
    current_dur = None
    current_wi = None

    for r in range(start_row, start_row + max_rows):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 19)]

        if _is_total_row(row_vals):
            return materials, r + 1

        if not _is_material_row(row_vals):
            if all(_clean(v) is None for v in row_vals):
                continue
            continue

        # Update process parameter jika ada isian baru
        temp = _clean(row_vals[MATERIAL_COL_TEMP - 1])
        agit = _clean(row_vals[MATERIAL_COL_AGIT - 1])
        dur  = _clean(row_vals[MATERIAL_COL_MIX_DUR - 1])
        wi   = _clean(row_vals[MATERIAL_COL_WI - 1])

        if temp is not None:
            current_temp = temp
        if agit is not None:
            current_agit = agit
        if dur is not None:
            current_dur = dur
        if wi is not None:
            current_wi = wi

        ratio_raw   = _clean_ratio(row_vals[MATERIAL_COL_RATIO - 1])
        dosage_raw  = _clean(row_vals[MATERIAL_COL_DOSAGE - 1])   # mungkin #REF!
        dosage_kg   = _clean_ratio(row_vals[MATERIAL_COL_DOSAGE_KG - 1])
        stick_per_mc = 10000  # default, bisa diupdate dari product header jika perlu

        # Hitung dosage_mg_stick jika raw-nya #REF! tapi ratio + product_weight tersedia
        if dosage_raw is None and ratio_raw is not None and product_weight_mg_stick is not None:
            dosage_mg_stick_calc = compute_dosage_mg_stick(
                ratio_raw, product_weight_mg_stick, phase_app_pct, phase_blend_ratio
            )
        else:
            dosage_mg_stick_calc = dosage_raw  # gunakan nilai asli jika ada

        # Hitung dosage_kg_mc jika belum ada
        if dosage_kg is None and dosage_mg_stick_calc is not None:
            dosage_kg = round(dosage_mg_stick_calc * stick_per_mc / 1_000_000, 9)

        ratio_pct = round(ratio_raw * 100, 6) if ratio_raw is not None else None

        mat: Dict[str, Any] = {
            "phase": phase_name,
            "no": _clean(row_vals[MATERIAL_COL_NO - 1]),
            "item_code": _clean(row_vals[MATERIAL_COL_ITEM_CODE - 1]),
            "item_name": _clean(row_vals[MATERIAL_COL_ITEM_NAME - 1]),
            "physical_form": _clean(row_vals[MATERIAL_COL_PHYSICAL_FORM - 1]),
            "cas_number": _clean(row_vals[MATERIAL_COL_CAS - 1]),
            "ratio_pct": ratio_pct,
            "dosage_mg_stick": dosage_mg_stick_calc,
            "material_price_usd_kg": _clean_ratio(row_vals[MATERIAL_COL_MAT_PRICE - 1]),
            "formulation_price_usd_kg": _clean_ratio(row_vals[MATERIAL_COL_FORM_PRICE - 1]),
            "dosage_kg_mc": dosage_kg,
            "density": _clean(row_vals[MATERIAL_COL_DENSITY - 1]),
            "addition_sequence": _clean(row_vals[MATERIAL_COL_SEQ - 1]),
            "temperature": current_temp,
            "agitation_rate": current_agit,
            "mixing_duration": current_dur,
            "work_instruction": current_wi,
        }
        materials.append(mat)

    return materials, start_row + max_rows


# ---------------------------------------------------------------------------
# Scraper Utama
# ---------------------------------------------------------------------------

def scrape_template(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
    """
    Scrape sheet Excel template dan ekstrak:
    - Header product info
    - Phase metadata (Casing Rajangan, Casing Krosok, Top Flavor, Premix-premix)
    - Material rows per phase
    - Summary casing & top flavor
    - Approval block
    """
    result: Dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "model_type": "formulation_product_scraped",
        "source_sheet": ws.title,
        "product": {},
        "phase_metadata": [],
        "materials": [],
        "summary": [],
        "approval": {},
    }

    def cell(r, c):
        return _clean(ws.cell(row=r, column=c).value)

    # -----------------------------------------------------------------------
    # 1. Header Product Information
    # -----------------------------------------------------------------------
    product: Dict[str, Any] = {}

    # Baca seluruh area header secara dinamis
    for r in range(1, 20):
        col1 = _clean(ws.cell(row=r, column=1).value)
        col3 = _clean(ws.cell(row=r, column=3).value)
        col5 = _clean(ws.cell(row=r, column=5).value)
        col9 = _clean(ws.cell(row=r, column=9).value)
        col10 = _clean(ws.cell(row=r, column=10).value)

        if col1 is None:
            continue

        label = str(col1).strip().lower()

        if "full fg description" in label or "product name" in label:
            product["product_name"] = col3
        elif "mixing factory" in label:
            product["mixing_factory"] = col3
        elif "lab personnel" in label:
            product["prepared_by"] = col3
        elif "standard control" in label:
            product["standard_control"] = col3
            product["flavor_standard_reference"] = col10 or _clean(ws.cell(row=r, column=9).value)
        elif "tobacco blend code" in label:
            product["tobacco_blend_code"] = col3
            product["sensory_parameter"] = col10
        elif "formulation code" in label:
            product["formulation_code"] = col3
            product["impact"] = col10
        elif "single capsule" in label:
            product["single_capsule"] = col3
            product["flavor_aroma"] = col10
        elif "double capsule (tobacco end)" in label:
            product["double_capsule_tobacco_end"] = col3
            product["irritation"] = col10
        elif "double capsule (mouth end)" in label:
            product["double_capsule_mouth_end"] = col3
            product["cooling"] = col10
        elif "product weight" in label or ("mg" in label and "stick" in label):
            product["product_weight_mg_stick"] = _clean_ratio(ws.cell(row=r, column=3).value)
            product["clove_weight_mg_stick"] = _clean_ratio(ws.cell(row=r, column=6).value)
        elif "no. of stick" in label or "stick per mc" in label:
            product["stick_per_mc"] = _clean_ratio(ws.cell(row=r, column=10).value)
        elif "revision no" in label and r <= 5:
            product["revision_no"] = _clean(ws.cell(row=r, column=17).value)
        elif "effective date" in label and r <= 5:
            product["effective_date"] = _clean(ws.cell(row=r, column=17).value)
        elif "document no" in label and r <= 5:
            product["document_no"] = _clean(ws.cell(row=r, column=17).value)

    # Cari Stick per MC jika belum ditemukan (bisa ada di row berbeda)
    if "stick_per_mc" not in product:
        for r in range(1, 25):
            for c in range(1, 18):
                v = _clean(ws.cell(row=r, column=c).value)
                if isinstance(v, (int, float)) and v == 10000:
                    product["stick_per_mc"] = int(v)
                    break

    result["product"] = product

    # -----------------------------------------------------------------------
    # 2. Deteksi dan Baca Phase Sections
    # -----------------------------------------------------------------------
    # Strategi: scan baris dan deteksi label section berdasarkan isi kolom 1
    phase_sections_found: List[Dict[str, Any]] = []  # {row, label}

    KNOWN_PHASE_LABELS = {
        "casing rajangan": "Casing Rajangan",
        "casing krosok": "Casing Krosok",
        "top flavor": "Top Flavor",
        "casing pre-mix": "Casing Pre-Mix",
        "flavor pre-mix": "Flavor Pre-Mix",
    }

    # Scan semua baris untuk menemukan section headers
    section_rows: List[Tuple[int, str]] = []
    for r in range(1, ws.max_row + 1):
        col1 = ws.cell(row=r, column=1).value
        if col1 is None:
            continue
        col1_str = str(col1).strip().lower()
        for key, label in KNOWN_PHASE_LABELS.items():
            if col1_str == key:
                section_rows.append((r, label))
                break
        # Deteksi premix dari struktur berbeda: "Casing Pre-Mix Formulation" atau NAV code di kolom 5
        if "casing pre-mix formulation" == col1_str:
            section_rows.append((r, "Casing Pre-Mix"))
        elif "flavor pre-mix formulation" == col1_str:
            # Dapatkan NAV Item Code untuk label premix yang spesifik
            nav_code = _clean(ws.cell(row=r + 1, column=5).value)
            nav_desc = _clean(ws.cell(row=r + 2, column=5).value)
            pm_label = nav_desc or nav_code or "Flavor Pre-Mix"
            section_rows.append((r, f"Flavor Pre-Mix [{pm_label}]"))

    # Hapus duplikat dan urutkan
    seen = set()
    unique_sections: List[Tuple[int, str]] = []
    for row_num, label in section_rows:
        if row_num not in seen:
            seen.add(row_num)
            unique_sections.append((row_num, label))
    unique_sections.sort(key=lambda x: x[0])

    # -----------------------------------------------------------------------
    # 3. Baca metadata dan material setiap section
    # -----------------------------------------------------------------------
    all_phase_metadata: List[Dict[str, Any]] = []
    all_materials: List[Dict[str, Any]] = []

    for sec_idx, (sec_row, sec_label) in enumerate(unique_sections):
        # Baca metadata block (NAV code, description, blend ratio, application)
        pm, mat_start_row = _parse_phase_metadata_block(ws, sec_row + 1, sec_label)
        all_phase_metadata.append(pm)

        # Tentukan batas akhir section ini
        end_limit = unique_sections[sec_idx + 1][0] if sec_idx + 1 < len(unique_sections) else ws.max_row

        # Baca material rows
        mats, _ = _read_material_block(ws, mat_start_row, sec_label, max_rows=end_limit - mat_start_row + 5)
        all_materials.extend(mats)

    result["phase_metadata"] = all_phase_metadata
    result["materials"] = all_materials

    # -----------------------------------------------------------------------
    # 4. Baca Summary Casing & Top Flavor
    # -----------------------------------------------------------------------
    summary_rows: List[Dict[str, Any]] = []
    for r in range(1, ws.max_row + 1):
        col1 = _clean(ws.cell(row=r, column=1).value)
        col2 = _clean(ws.cell(row=r, column=2).value)
        if col1 is None and col2 is None:
            continue
        # Baris summary: kolom 1 adalah nomor (1, 2, 3) dan kolom 2 berisi kode formula
        col1_v = col1
        col4 = _clean_ratio(ws.cell(row=r, column=4).value)
        col7 = _clean_ratio(ws.cell(row=r, column=7).value)
        col8 = _clean_ratio(ws.cell(row=r, column=8).value)

        if col1_v in (1, 2, 3) and col2 and col4 is not None:
            # Kemungkinan baris summary
            summary_rows.append({
                "no": col1_v,
                "nav_item_description": col2,
                "blend_ratio": round(col4 * 100, 4) if col4 else None,
                "price_usd_kg": col7,
                "price_1000_sticks": col8,
            })

    # Ambil hanya 3 baris pertama jika lebih
    if summary_rows:
        result["summary"] = summary_rows[:3]

    # -----------------------------------------------------------------------
    # 5. Baca Approval Block
    # -----------------------------------------------------------------------
    approval: Dict[str, Any] = {}
    for r in range(1, ws.max_row + 1):
        col1 = _clean(ws.cell(row=r, column=1).value)
        col6 = _clean(ws.cell(row=r, column=6).value)
        col13 = _clean(ws.cell(row=r, column=13).value)

        if col1 is None:
            continue
        col1_str = str(col1).strip().lower()

        if col1_str == "prepared by":
            approval["prepared_by_label"] = col1
            approval["reviewed_by_label"] = col6
            approval["approved_by_label"] = col13
        elif col1_str.startswith("position"):
            approval["prepared_position"] = col1
            approval["reviewed_position"] = col6
            approval["approved_position"] = col13
        elif col1_str.startswith("name"):
            approval["prepared_name"] = col1
            approval["reviewed_name"] = col6
            approval["approved_name"] = col13
        elif col1_str.startswith("date"):
            approval["prepared_date"] = col1
            approval["reviewed_date"] = col6
            approval["approved_date"] = col13

    result["approval"] = approval

    return result


# ---------------------------------------------------------------------------
# Export ke JSON
# ---------------------------------------------------------------------------

def _serialize_value(v: Any) -> Any:
    if isinstance(v, (datetime, date_class)):
        return str(v)
    if isinstance(v, float) and (v != v):
        return None
    return v


def export_to_json(data: Dict[str, Any]) -> bytes:
    def default_serializer(obj):
        if isinstance(obj, (datetime, date_class)):
            return str(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

    return json.dumps(data, indent=2, ensure_ascii=False, default=default_serializer).encode("utf-8")


# ---------------------------------------------------------------------------
# Export ke XLSX normalisasi (3 sheet: product, phase_metadata, materials)
# ---------------------------------------------------------------------------

def export_to_xlsx(data: Dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Sheet 1: product
        product_row = {k: str(v) if isinstance(v, (datetime, date_class)) else v
                       for k, v in data.get("product", {}).items()}
        product_row["source_sheet"] = data.get("source_sheet", "")
        pd.DataFrame([product_row]).to_excel(writer, sheet_name="product", index=False)

        # Sheet 2: phase_metadata
        phase_rows = data.get("phase_metadata", [])
        if phase_rows:
            pd.DataFrame(phase_rows).to_excel(writer, sheet_name="phase_metadata", index=False)
        else:
            pd.DataFrame(columns=["phase", "nav_code", "description", "blend_ratio", "application"]).to_excel(
                writer, sheet_name="phase_metadata", index=False
            )

        # Sheet 3: materials
        mat_rows = data.get("materials", [])
        if mat_rows:
            pd.DataFrame(mat_rows).to_excel(writer, sheet_name="materials", index=False)
        else:
            pd.DataFrame(columns=["phase", "no", "item_code", "item_name"]).to_excel(
                writer, sheet_name="materials", index=False
            )

        # Sheet 4: summary
        summary_rows = data.get("summary", [])
        if summary_rows:
            pd.DataFrame(summary_rows).to_excel(writer, sheet_name="summary", index=False)

        # Sheet 5: approval
        approval = data.get("approval", {})
        if approval:
            pd.DataFrame([approval]).to_excel(writer, sheet_name="approval", index=False)

    buffer.seek(0)
    return buffer.read()


# ---------------------------------------------------------------------------
# Streamlit App
# ---------------------------------------------------------------------------

def render_template_scraper_app(*, embedded: bool = False) -> None:
    if not embedded:
        st.set_page_config(
            page_title="📋 Template Scraper — Formulation Sheet",
            page_icon="⚗️",
            layout="wide",
        )

    st.title("Data Extraction Engine")
    st.markdown(
        "Upload file Excel template formulasi (format asli / legacy seperti `template_1.xlsx`). "
        "Sistem akan otomatis membaca data produk, fase, material, dan approval block, "
        "lalu menyimpannya ke JSON dan XLSX normalisasi."
    )

    default_template = PROJECT_ROOT / "template_1.xlsx"

    use_default = False
    if default_template.exists():
        use_default = st.checkbox(
            f"Gunakan file default `{default_template.name}` yang sudah ada di folder proyek",
            value=True,
            key="scraper_use_default_template",
        )

    uploaded_file = None
    if not use_default:
        uploaded_file = st.file_uploader(
            "Upload file Excel template (.xlsx / .xlsm)",
            type=["xlsx", "xlsm"],
            key="scraper_uploaded_template",
        )

    source_bytes: Optional[bytes] = None
    source_name = ""

    if use_default and default_template.exists():
        source_bytes = default_template.read_bytes()
        source_name = default_template.name
    elif uploaded_file is not None:
        source_bytes = uploaded_file.read()
        source_name = uploaded_file.name

    if source_bytes is None:
        st.info("Upload file Excel atau centang opsi default untuk memulai.")
        return

    st.divider()
    st.subheader("Pilih Sheet")

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(source_bytes), data_only=True)
    except Exception as exc:
        st.error(f"Gagal membuka file: {exc}")
        return

    selected_sheet = st.selectbox(
        "Sheet yang akan di-scrape:",
        options=workbook.sheetnames,
        index=0,
        key="scraper_selected_sheet",
    )
    worksheet = workbook[selected_sheet]

    st.caption(
        f"File: `{source_name}` | Sheet: `{selected_sheet}` | "
        f"Total rows: {worksheet.max_row} | Total cols: {worksheet.max_column}"
    )

    st.divider()
    with st.spinner("Sedang membaca dan mengekstrak data template..."):
        scraped_data = scrape_template(worksheet)

    product_info = scraped_data.get("product", {})
    phases = scraped_data.get("phase_metadata", [])
    materials = scraped_data.get("materials", [])
    summary = scraped_data.get("summary", [])
    approval = scraped_data.get("approval", {})

    st.success(
        f"Scraping selesai: ditemukan {len(phases)} fase dan {len(materials)} material rows."
    )

    with st.expander("Product Header Info", expanded=True):
        col_a, col_b = st.columns(2)
        with col_a:
            for key in [
                "product_name",
                "mixing_factory",
                "prepared_by",
                "formulation_code",
                "tobacco_blend_code",
            ]:
                st.markdown(f"**{key.replace('_', ' ').title()}**: {product_info.get(key, '—')}")
        with col_b:
            for key in [
                "product_weight_mg_stick",
                "clove_weight_mg_stick",
                "stick_per_mc",
                "standard_control",
                "revision_no",
                "effective_date",
            ]:
                st.markdown(f"**{key.replace('_', ' ').title()}**: {product_info.get(key, '—')}")

    with st.expander("Phase Metadata", expanded=True):
        if phases:
            st.dataframe(pd.DataFrame(phases), use_container_width=True, hide_index=True)
        else:
            st.warning("Tidak ada data fase ditemukan.")

    with st.expander("Material Rows", expanded=False):
        if materials:
            df_mat = pd.DataFrame(materials)
            st.dataframe(df_mat, use_container_width=True, hide_index=True)
            st.caption(f"Total: {len(df_mat)} material rows dari {df_mat['phase'].nunique()} fase.")
        else:
            st.warning("Tidak ada material ditemukan.")

    with st.expander("Summary Casing & Top Flavor", expanded=False):
        if summary:
            st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)
        else:
            st.info("Tidak ada data summary ditemukan.")

    with st.expander("Approval Block", expanded=False):
        if approval:
            st.json(approval)
        else:
            st.info("Tidak ada approval block ditemukan.")

    st.divider()
    st.subheader("Simpan ke Database")

    safe_name = re.sub(r"[^\w\-_.]", "_", source_name.replace(".xlsx", "").replace(".xlsm", ""))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_filename = f"scraped_{safe_name}_{timestamp}"

    json_bytes = export_to_json(scraped_data)
    xlsx_bytes = export_to_xlsx(scraped_data)

    json_output_path = OUTPUT_DIR / f"{base_filename}.json"
    xlsx_output_path = OUTPUT_DIR / f"{base_filename}.xlsx"
    json_output_path.write_bytes(json_bytes)
    xlsx_output_path.write_bytes(xlsx_bytes)

    st.success("File berhasil disimpan ke folder `outputs/scraped/`.")

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="Download JSON Database",
            data=json_bytes,
            file_name=f"{base_filename}.json",
            mime="application/json",
            use_container_width=True,
        )
    with col2:
        st.download_button(
            label="Download XLSX Database",
            data=xlsx_bytes,
            file_name=f"{base_filename}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


if __name__ == "__main__":
    render_template_scraper_app()
