import unittest
import json
import csv
import io
import zipfile
from copy import deepcopy
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from formulation_generator import (
    DEFAULT_APPROVED_BY,
    DEFAULT_APPROVED_POSITION,
    DEFAULT_EFFECTIVE_DATE,
    DEFAULT_PREPARED_BY_NAME,
    DEFAULT_PREPARED_POSITION,
    DEFAULT_REVIEWED_BY,
    DEFAULT_REVIEWED_POSITION,
    FORMULA_ROW_HEIGHT_POINTS,
    FORMULA_MATERIAL_COLUMNS,
    INPUT_SOURCE_FONT_COLOR,
    MATERIAL_DB_PATH,
    PARENT_BOM_LEVEL1_FILL_COLOR,
    PHASE_PERCENT_FORMAT,
    SECTION_ROW_RANGES,
    USD_ACCOUNTING_FORMAT,
    build_formulation_code_formula,
    build_input_formula_code_formula,
    build_input_formulation_code_formula,
    build_formulation_input_from_dict,
    create_blank_input_workbook,
    export_formulation_model_json,
    export_formulation_model_xlsx,
    export_formulation_model_csv_bundle,
    generate_formulation_workbook,
    set_cell_value,
)


def font_rgb(cell):
    color = cell.font.color
    return getattr(color, "rgb", None)


def fill_rgb(cell):
    return getattr(cell.fill.fgColor, "rgb", None)


class TestFormulationGenerator(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.root_dir = Path(__file__).resolve().parents[1]
        cls.template_path = cls.root_dir / "templates" / "Template_Generate.xlsm"
        if not cls.template_path.exists():
            raise FileNotFoundError(f"Test template not found: {cls.template_path}")

        cls.form_data = {
            "product_name": "SAAT 369",
            "formula_code": "CS_K-CK-LCDT-04_R00",
            "product_weight_mg_stick": 720,
            "clove_weight_mg_stick": 240,
            "stick_per_mc": 10000,
            "prepared_by": "Lyla Isro",
            "date": "2026-07-07",
            "approval_date": "2026-07-09",
            "standard_control": "DJARUM 76",
            "flavor_standard_reference": "N/A",
            "tobacco_blend_code": "SO-18",
            "sensory_parameter": "Sample",
            "formulation_code": "CS_K-CK-LCDT-04_R00",
            "impact": "Low",
            "single_capsule": "N/A",
            "flavor_aroma": "Spice",
            "double_capsule_tobacco_end": "N/A",
            "irritation": "None",
            "double_capsule_mouth_end": "N/A",
            "cooling": "Mild",
            "phase_metadata": {
                "Casing Rajangan": {
                    "nav_code": "GEN-CS-00000",
                    "description": "CS_R-CK-LCDT-04_R00",
                    "blend_ratio": 0.38,
                    "application": 0.03078,
                },
            },
        }
        cls.material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.25,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
        ]

    def build_formulation(self, form_data=None, material_rows=None):
        return build_formulation_input_from_dict(
            deepcopy(form_data if form_data is not None else self.form_data),
            deepcopy(material_rows if material_rows is not None else self.material_rows),
        )

    def with_required_main_phases(self, material_rows):
        rows = deepcopy(material_rows)
        existing_phases = {row.get("phase") for row in rows}
        required_defaults = {
            "Casing Rajangan": {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            "Casing Krosok": {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            "Top Flavor": {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
        }
        for phase, row in required_defaults.items():
            if phase not in existing_phases:
                rows.append(row)
        return rows

    def generate_workbook(self, form_data=None, material_rows=None):
        formulation = self.build_formulation(form_data, material_rows)
        temp_dir = TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        output_path = Path(temp_dir.name) / "formulation_test.xlsx"
        generate_formulation_workbook(formulation, self.template_path, output_path)
        return openpyxl.load_workbook(output_path, data_only=False)

    def assert_formula_or_value(self, actual, expected, formula_prefix=None):
        if isinstance(actual, str) and actual.startswith("="):
            if formula_prefix is not None:
                self.assertTrue(
                    actual.startswith(formula_prefix),
                    f"Expected formula beginning with {formula_prefix!r}, got {actual!r}",
                )
            return
        self.assertAlmostEqual(actual, expected)

    def test_generate_formulation_workbook_creates_xlsx(self):
        self.assertTrue(MATERIAL_DB_PATH.exists())
        formulation = self.build_formulation()
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "generated" / "formulation_test.xlsx"
            result_path = generate_formulation_workbook(formulation, self.template_path, output_path)

            self.assertTrue(result_path.exists())
            self.assertEqual(result_path, output_path)

            workbook = openpyxl.load_workbook(result_path)
            self.assertIn("INPUT", workbook.sheetnames)
            self.assertIn("Formula", workbook.sheetnames)
            self.assertEqual(workbook["Formula"].row_dimensions[1].height, 20.0)
            self.assertEqual(workbook["Formula"].row_dimensions[2].height, 20.0)
            self.assertEqual(workbook["Formula"].row_dimensions[3].height, 20.0)
            self.assertEqual(workbook["Formula"].row_dimensions[25].height, 48.0)
            self.assertEqual(workbook["Formula"].row_dimensions[26].height, FORMULA_ROW_HEIGHT_POINTS)
            self.assertEqual(workbook["INPUT"].cell(row=2, column=4).value, "PHASE / CASING / FLAVOR METADATA")
            self.assertEqual(workbook["INPUT"].cell(row=4, column=5).value, "GEN-CS-00000")
            self.assertEqual(workbook["INPUT"].cell(row=4, column=6).value, "CS_R-CK-LCDT-04_R00")
            self.assertEqual(workbook["INPUT"].cell(row=4, column=7).value, 0.38)
            self.assertEqual(workbook["INPUT"].cell(row=4, column=8).value, 0.03078)
            self.assertEqual(workbook["INPUT"].cell(row=4, column=7).number_format, PHASE_PERCENT_FORMAT)
            self.assertEqual(workbook["INPUT"].cell(row=4, column=8).number_format, PHASE_PERCENT_FORMAT)
            self.assertEqual(workbook["Formula"].cell(row=23, column=5).number_format, PHASE_PERCENT_FORMAT)
            self.assertEqual(workbook["Formula"].cell(row=24, column=5).number_format, PHASE_PERCENT_FORMAT)
            self.assertEqual(workbook["BOL-SAAT List"].cell(row=1, column=6).value, "Chemical Name")
            self.assertEqual(workbook["BOL-SAAT List"].cell(row=2, column=1).value, "GEN-AC-00003")
            self.assertAlmostEqual(workbook["BOL-SAAT List"].cell(row=2, column=7).value, 84.36)

            with zipfile.ZipFile(result_path) as workbook_zip:
                self.assertFalse(
                    any(name.startswith("xl/externalLinks/") for name in workbook_zip.namelist()),
                    "Generated .xlsx should not preserve external links from the source template.",
                )
                self.assertTrue(
                    any(name.startswith("xl/media/") for name in workbook_zip.namelist()),
                    "Generated .xlsx should include the project logo image.",
                )

    def test_export_formulation_model_json_has_database_like_structure(self):
        formulation = self.build_formulation()
        payload = json.loads(export_formulation_model_json(formulation).decode("utf-8"))

        self.assertEqual(payload["schema_version"], "1.0")
        self.assertEqual(payload["model_type"], "formulation_product")
        self.assertIn("product", payload)
        self.assertIn("phase_metadata", payload)
        self.assertIn("materials", payload)
        self.assertEqual(payload["product"]["formula_code"], "CS_K-CK-LCDT-04_R00")
        self.assertIsInstance(payload["phase_metadata"], list)
        self.assertIsInstance(payload["materials"], list)
        self.assertGreaterEqual(len(payload["materials"]), 1)
        self.assertEqual(payload["materials"][0]["phase"], "Casing Rajangan")

    def test_build_formulation_input_treats_nan_optional_material_fields_as_blank(self):
        rows = deepcopy(self.material_rows)
        rows[0]["addition_sequence"] = float("nan")
        rows[0]["temperature"] = float("nan")

        formulation = self.build_formulation(material_rows=rows)

        self.assertIsNone(formulation.materials[0].addition_sequence)
        self.assertIsNone(formulation.materials[0].temperature)

    def test_export_formulation_model_csv_bundle_contains_three_tables(self):
        formulation = self.build_formulation()
        bundle = export_formulation_model_csv_bundle(formulation)

        with zipfile.ZipFile(io.BytesIO(bundle)) as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                ["materials.csv", "phase_metadata.csv", "product.csv"],
            )
            with archive.open("product.csv") as handle:
                product_rows = list(csv.DictReader(io.TextIOWrapper(handle, encoding="utf-8")))
            with archive.open("phase_metadata.csv") as handle:
                phase_rows = list(csv.DictReader(io.TextIOWrapper(handle, encoding="utf-8")))
            with archive.open("materials.csv") as handle:
                material_rows = list(csv.DictReader(io.TextIOWrapper(handle, encoding="utf-8")))

        self.assertEqual(product_rows[0]["formula_code"], "CS_K-CK-LCDT-04_R00")
        self.assertGreaterEqual(len(phase_rows), 1)
        self.assertGreaterEqual(len(material_rows), 1)
        self.assertEqual(material_rows[0]["phase"], "Casing Rajangan")

    def test_export_formulation_model_xlsx_contains_three_sheets(self):
        formulation = self.build_formulation()
        workbook = openpyxl.load_workbook(io.BytesIO(export_formulation_model_xlsx(formulation)))

        self.assertEqual(workbook.sheetnames, ["product", "phase_metadata", "materials"])
        self.assertEqual(workbook["product"]["A1"].value, "product_id")
        self.assertEqual(workbook["product"]["B1"].value, "product_name")
        self.assertEqual(workbook["product"]["C1"].value, "formula_code")
        self.assertEqual(workbook["product"]["A2"].value, "CS-K-CK-LCDT-04-R00")
        self.assertEqual(workbook["product"]["I1"].value, "approval_date")
        self.assertEqual(workbook["product"]["I2"].value, "2026-07-09")
        self.assertEqual(workbook["phase_metadata"]["A1"].value, "phase_id")
        self.assertEqual(workbook["phase_metadata"]["B1"].value, "product_id")
        self.assertEqual(workbook["phase_metadata"]["A2"].value, "CS-K-CK-LCDT-04-R00::CASING-RAJANGAN")
        self.assertEqual(workbook["materials"]["A1"].value, "material_id")
        self.assertEqual(workbook["materials"]["B1"].value, "product_id")
        self.assertEqual(workbook["materials"]["C1"].value, "phase_id")
        self.assertEqual(workbook["materials"]["D2"].value, "Casing Rajangan")

    def test_generated_input_sheet_contains_process_columns(self):
        material_rows = deepcopy(self.material_rows)
        material_rows[0].update({
            "temperature": "45",
            "agitation_rate": "500 RPM",
            "mixing_duration": "15 min",
            "work_instruction_override": "Start mixing base materials.",
            "process_role": "Carrier/Base",
            "notes": "Lab note",
        })
        formulation = self.build_formulation(material_rows=material_rows)
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "input_process_columns.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            input_sheet = workbook["INPUT"]

            expected_headers = {
                "G19": "Addition Sequence",
                "H19": "Temperature",
                "I19": "Agitation Rate",
                "J19": "Mixing Duration",
                "K19": "Work Instruction Override",
                "L19": "Process Role",
                "M19": "Notes",
            }
            for cell_ref, expected_header in expected_headers.items():
                self.assertEqual(input_sheet[cell_ref].value, expected_header)

            self.assertEqual(input_sheet["G20"].value, 1)
            self.assertEqual(input_sheet["H20"].value, "45 °C")
            self.assertEqual(input_sheet["I20"].value, "500 RPM")
            self.assertEqual(input_sheet["J20"].value, "15 min")
            self.assertEqual(input_sheet["K20"].value, "Start mixing base materials.")
            self.assertEqual(input_sheet["L20"].value, "Carrier/Base")
            self.assertEqual(input_sheet["M20"].value, "Lab note")

    def test_generated_workbook_contains_metadata(self):
        formulation = self.build_formulation()
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "formulation_test.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path)

            formula_sheet = workbook["Formula"]
            self.assertIn("A1:B3", [str(cell_range) for cell_range in formula_sheet.merged_cells.ranges])
            self.assertIsNone(formula_sheet["A1"].value)
            self.assertEqual(formula_sheet.cell(row=5, column=3).value, "SAAT 369")
            self.assertEqual(formula_sheet.cell(row=6, column=1).value, "Mixing Factory")
            self.assertEqual(formula_sheet.cell(row=6, column=3).value, "PT SAAT")
            self.assertEqual(formula_sheet.cell(row=7, column=3).value, "Lyla Isro")
            self.assertEqual(formula_sheet.cell(row=1, column=16).value, "Document No")
            self.assertEqual(formula_sheet.cell(row=1, column=17).value, "BOL-ID-BOLL-SPEC-003")
            self.assertEqual(formula_sheet.cell(row=2, column=16).value, "Revision No")
            self.assertEqual(formula_sheet.cell(row=2, column=17).value, "00")
            self.assertEqual(formula_sheet.cell(row=3, column=16).value, "Effective Date")
            self.assertEqual(formula_sheet.cell(row=3, column=17).value, "2026-07-07")
            self.assertEqual(formula_sheet.cell(row=10, column=8).value, "Flavor Standard Reference")
            self.assertEqual(formula_sheet.cell(row=10, column=10).value, "N/A")
            self.assertEqual(formula_sheet.cell(row=12, column=8).value, "Sensory Parameter")
            self.assertEqual(formula_sheet.cell(row=13, column=8).value, "Impact")
            self.assertEqual(formula_sheet.cell(row=14, column=8).value, "Flavor Aroma")
            self.assertEqual(formula_sheet.cell(row=15, column=8).value, "Irritation")
            self.assertEqual(formula_sheet.cell(row=16, column=8).value, "Cooling")
            self.assertEqual(formula_sheet.cell(row=21, column=5).value, "GEN-CS-00000")
            self.assertEqual(formula_sheet.cell(row=22, column=5).value, "CS_R-CK-LCDT-04_R00")
            self.assertEqual(formula_sheet.cell(row=13, column=3).value, build_formulation_code_formula())

    def test_missing_effective_date_defaults_to_template_release_date(self):
        form_data = deepcopy(self.form_data)
        form_data["date"] = ""
        workbook = self.generate_workbook(form_data=form_data)
        formula_sheet = workbook["Formula"]

        self.assertEqual(formula_sheet.cell(row=3, column=17).value, DEFAULT_EFFECTIVE_DATE.isoformat())

    def test_casing_top_flavor_summary_uses_phase_description_totals_and_usd_price(self):
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        header_row = None
        for row in range(1, formula_sheet.max_row + 1):
            value = formula_sheet.cell(row=row, column=2).value
            if isinstance(value, str) and "Casing & Top Flavor" in value:
                header_row = row
                break

        self.assertIsNotNone(header_row)
        first_summary_row = header_row + 1
        first_total_row = SECTION_ROW_RANGES["Casing Rajangan"][-1] + 1

        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=2).value, "=E22")
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=3).value, f"=H{first_total_row}")
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=4).value, "=E23")
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=7).value, f"=J{first_total_row}")
        self.assertEqual(
            formula_sheet.cell(row=first_summary_row, column=8).value,
            f"=G{first_summary_row}*(C{first_summary_row}/1000000)*1000",
        )
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=7).number_format, USD_ACCOUNTING_FORMAT)
        self.assertEqual(formula_sheet.cell(row=first_summary_row, column=8).number_format, USD_ACCOUNTING_FORMAT)

    def test_approval_block_contains_prepared_reviewed_and_approved_columns(self):
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        approval_row = None
        for row in range(1, formula_sheet.max_row + 1):
            if formula_sheet.cell(row=row, column=1).value == "Prepared By":
                approval_row = row
                break

        self.assertIsNotNone(approval_row)
        self.assertEqual(formula_sheet.cell(row=approval_row, column=1).value, "Prepared By")
        self.assertEqual(formula_sheet.cell(row=approval_row, column=6).value, "Reviewed By")
        self.assertEqual(formula_sheet.cell(row=approval_row, column=13).value, "Approved By")
        merged_ranges = {str(cell_range) for cell_range in formula_sheet.merged_cells.ranges}
        self.assertIn(f"A{approval_row}:E{approval_row}", merged_ranges)
        self.assertIn(f"F{approval_row}:L{approval_row}", merged_ranges)
        self.assertIn(f"M{approval_row}:Q{approval_row}", merged_ranges)
        self.assertIn(f"A{approval_row + 1}:E{approval_row + 1}", merged_ranges)
        self.assertIn(f"F{approval_row + 1}:L{approval_row + 1}", merged_ranges)
        self.assertIn(f"M{approval_row + 1}:Q{approval_row + 1}", merged_ranges)
        self.assertIn(f"A{approval_row + 2}:E{approval_row + 2}", merged_ranges)
        self.assertIn(f"F{approval_row + 2}:L{approval_row + 2}", merged_ranges)
        self.assertIn(f"M{approval_row + 2}:Q{approval_row + 2}", merged_ranges)
        self.assertEqual(formula_sheet.cell(row=approval_row + 1, column=1).value, "Position : Flavourist\nName : Lyla Isro\nDate : 09 July 2026")
        self.assertEqual(formula_sheet.cell(row=approval_row + 1, column=6).value, "Position : Senior Manager - Flavourist\nName : Mochamad Setyawan\nDate : 09 July 2026")
        self.assertEqual(formula_sheet.cell(row=approval_row + 1, column=13).value, "Position : Head, Flavour PDI\nName : Andrew Yip\nDate : 09 July 2026")
        self.assertEqual(formula_sheet.cell(row=approval_row + 2, column=1).value, "Signature:")
        self.assertEqual(formula_sheet.row_dimensions[approval_row].height, 18.0)
        self.assertEqual(formula_sheet.row_dimensions[approval_row + 1].height, 48.0)
        self.assertEqual(formula_sheet.row_dimensions[approval_row + 2].height, 54.0)

    def test_blank_input_workbook_can_be_downloaded_as_xlsx(self):
        workbook_bytes = create_blank_input_workbook(self.template_path)
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "blank_input.xlsx"
            output_path.write_bytes(workbook_bytes)
            workbook = openpyxl.load_workbook(output_path)

        self.assertEqual(workbook.sheetnames, ["INPUT"])
        input_sheet = workbook["INPUT"]
        self.assertEqual(input_sheet.row_dimensions[1].height, 20.0)
        self.assertEqual(input_sheet.row_dimensions[3].height, 20.0)
        self.assertIsNone(input_sheet.cell(row=2, column=2).value)
        self.assertEqual(input_sheet.cell(row=4, column=4).value, "Casing Rajangan")
        self.assertIsNone(input_sheet.cell(row=4, column=5).value)
        self.assertEqual(input_sheet.cell(row=3, column=2).value, build_input_formula_code_formula())
        self.assertEqual(input_sheet.cell(row=6, column=11).value, build_input_formulation_code_formula())
        self.assertEqual(input_sheet.cell(row=10, column=10).value, "APPROVAL METADATA")
        self.assertEqual(input_sheet.cell(row=7, column=1).value, "Lab Personnel Involved")
        self.assertEqual(input_sheet.cell(row=11, column=11).value, DEFAULT_PREPARED_BY_NAME)
        self.assertEqual(input_sheet.cell(row=12, column=11).value, DEFAULT_PREPARED_POSITION)
        self.assertEqual(input_sheet.cell(row=13, column=11).value, DEFAULT_REVIEWED_BY)
        self.assertEqual(input_sheet.cell(row=14, column=11).value, DEFAULT_REVIEWED_POSITION)
        self.assertEqual(input_sheet.cell(row=15, column=11).value, DEFAULT_APPROVED_BY)
        self.assertEqual(input_sheet.cell(row=16, column=11).value, DEFAULT_APPROVED_POSITION)
        self.assertEqual(input_sheet.cell(row=19, column=1).value, "Phase")
        self.assertEqual(input_sheet.cell(row=19, column=2).value, "Item Code")
        self.assertEqual(input_sheet.cell(row=19, column=3).value, "Item Name")
        self.assertIsNone(input_sheet.cell(row=20, column=2).value)
        self.assertEqual(input_sheet.cell(row=20, column=6).number_format, PHASE_PERCENT_FORMAT)
        self.assertTrue(input_sheet.protection.sheet)
        self.assertFalse(input_sheet.cell(row=4, column=5).protection.locked)
        self.assertTrue(input_sheet.cell(row=7, column=7).protection.locked)
        self.assertTrue(input_sheet.cell(row=7, column=8).protection.locked)
        self.assertFalse(input_sheet.cell(row=7, column=5).protection.locked)
        validation_ranges = {str(validation.sqref) for validation in input_sheet.data_validations.dataValidation}
        self.assertIn("A20:A269", validation_ranges)
        self.assertIn("D20:D269", validation_ranges)
        self.assertIn("L20:L269", validation_ranges)

    def test_input_sheet_separates_lab_personnel_and_prepared_by_name(self):
        form_data = deepcopy(self.form_data)
        form_data["prepared_by"] = "Wawan, Isro, Lyla"
        form_data["prepared_by_name"] = "Prepared User"
        material_rows = deepcopy(self.material_rows)
        material_rows[0]["temperature"] = "40 C"
        workbook = self.generate_workbook(form_data=form_data, material_rows=material_rows)
        input_sheet = workbook["INPUT"]

        self.assertEqual(input_sheet.cell(row=7, column=1).value, "Lab Personnel Involved")
        self.assertEqual(input_sheet.cell(row=7, column=2).value, "Wawan, Isro, Lyla")
        self.assertEqual(input_sheet.cell(row=11, column=10).value, "Prepared By Name")
        self.assertEqual(input_sheet.cell(row=11, column=11).value, "Prepared User")
        self.assertEqual(input_sheet.cell(row=20, column=2).value, "GEN-WA-00001")
        self.assertEqual(input_sheet.cell(row=20, column=3).value, "WATER")
        self.assertEqual(input_sheet.cell(row=20, column=8).value, "40 °C")
        self.assertEqual(input_sheet.cell(row=20, column=6).number_format, PHASE_PERCENT_FORMAT)

    def test_approval_block_uses_custom_people_and_positions(self):
        form_data = deepcopy(self.form_data)
        form_data["prepared_position"] = "Lead Flavourist"
        form_data["prepared_by_name"] = "Prepared User"
        form_data["reviewed_by"] = "Reviewer Name"
        form_data["reviewed_position"] = "Reviewer Position"
        form_data["approved_by"] = "Approver Name"
        form_data["approved_position"] = "Approver Position"

        workbook = self.generate_workbook(form_data=form_data)
        formula_sheet = workbook["Formula"]

        approval_row = next(
            row for row in range(1, formula_sheet.max_row + 1)
            if formula_sheet.cell(row=row, column=1).value == "Prepared By"
        )
        self.assertEqual(
            formula_sheet.cell(row=approval_row + 1, column=1).value,
            "Position : Lead Flavourist\nName : Prepared User\nDate : 09 July 2026",
        )
        self.assertEqual(
            formula_sheet.cell(row=approval_row + 1, column=6).value,
            "Position : Reviewer Position\nName : Reviewer Name\nDate : 09 July 2026",
        )
        self.assertEqual(
            formula_sheet.cell(row=approval_row + 1, column=13).value,
            "Position : Approver Position\nName : Approver Name\nDate : 09 July 2026",
        )

    def test_material_section_has_written_rows(self):
        formulation = self.build_formulation()
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "formulation_test.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path)

            formula_sheet = workbook["Formula"]
            first_row = formula_sheet.cell(row=26, column=2).value
            second_row = formula_sheet.cell(row=27, column=2).value
            self.assertEqual(first_row, "GEN-WA-00001")
            self.assertEqual(second_row, "GEN-CHM-00015")

    def test_material_sections_resize_to_written_materials_and_totals_match(self):
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        row_indices = SECTION_ROW_RANGES["Casing Rajangan"]
        self.assertEqual(len(row_indices), 2)
        total_row = row_indices[-1] + 1
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=7).value, 1.0, "=SUM(G26:")
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=8).value, 3.25, "=SUM(H26:")
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=10).value, 0.8492307692307692, "=SUM(J26:")
        self.assert_formula_or_value(formula_sheet.cell(row=total_row, column=11).value, 0.0325, "=SUM(K26:")
        self.assertIsNone(formula_sheet.cell(row=total_row, column=1).border.right.style)
        self.assertIsNone(formula_sheet.cell(row=total_row, column=2).border.left.style)
        self.assertEqual(formula_sheet.cell(row=total_row, column=1).border.top.style, "thin")
        self.assertEqual(formula_sheet.cell(row=total_row, column=17).border.right.style, "thin")
        self.assertEqual(formula_sheet.cell(row=total_row, column=7).border.left.style, "thin")
        self.assertEqual(formula_sheet.cell(row=total_row, column=7).border.right.style, "thin")
        self.assertEqual(formula_sheet.cell(row=total_row, column=17).border.left.style, "thin")

    def test_material_calculations_include_ratio_formulation_price_and_dosage_kg_mc(self):
        self.assertIn("[$$-409]", USD_ACCOUNTING_FORMAT)
        self.assertIn("#,##0.00", USD_ACCOUNTING_FORMAT)
        workbook = self.generate_workbook()
        formula_sheet = workbook["Formula"]

        total_row = SECTION_ROW_RANGES["Casing Rajangan"][-1] + 1
        expected_rows = {26: {}, 27: {}}

        for row, expected in expected_rows.items():
            self.assertEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["ratio"]).value,
                f"=IF($H${total_row}=0,0,H{row}/$H${total_row})",
            )
            self.assertEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["formulation_price"]).value,
                f"=G{row}*I{row}",
            )
            self.assertEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["material_price"]).number_format,
                USD_ACCOUNTING_FORMAT,
            )
            self.assertEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["formulation_price"]).number_format,
                USD_ACCOUNTING_FORMAT,
            )
            self.assertEqual(
                formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["dosage_kg_mc"]).value,
                f"=H{row}*$J$18/1000000",
            )
            self.assertTrue(
                font_rgb(formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"])).endswith(
                    INPUT_SOURCE_FONT_COLOR
                )
            )
            self.assertEqual(
                font_rgb(formula_sheet.cell(row=row, column=FORMULA_MATERIAL_COLUMNS["ratio"])),
                "00000000",
            )

    def test_percent_mode_converts_to_dosage_mg_stick_with_five_decimals(self):
        form_data = deepcopy(self.form_data)
        form_data["phase_metadata"]["Casing Krosok"] = {
            "nav_code": "GEN-CS-00000",
            "description": "CS_K-CK-LCDT-04_R00",
            "blend_ratio": 0.52,
            "application": 0.0283,
        }
        material_rows = self.with_required_main_phases([
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_input_mode": "%",
                "ratio_percent": 38,
                "application_percent": 3.078,
                "addition_sequence": 1,
            },
        ])
        workbook = self.generate_workbook(form_data=form_data, material_rows=material_rows)
        formula_sheet = workbook["Formula"]

        self.assertEqual(formula_sheet.cell(row=26, column=FORMULA_MATERIAL_COLUMNS["ratio"]).value, 0.38)
        self.assertTrue(
            font_rgb(formula_sheet.cell(row=26, column=FORMULA_MATERIAL_COLUMNS["ratio"])).endswith(
                INPUT_SOURCE_FONT_COLOR
            )
        )
        self.assertEqual(
            font_rgb(formula_sheet.cell(row=26, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"])),
            "00000000",
        )
        self.assertEqual(
            formula_sheet.cell(row=26, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).value,
            "=G26*($C$18-$F$18)*$E$24*($E$23/($E$23+$E$36))",
        )
        self.assertEqual(formula_sheet.cell(row=26, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).number_format, "0.00000")

    def test_percent_mode_uses_phase_metadata_application_without_material_application_column(self):
        form_data = deepcopy(self.form_data)
        form_data["phase_metadata"]["Top Flavor"] = {
            "nav_code": "TF-GA-09",
            "description": "TF-GA-09",
            "blend_ratio": 100,
            "application": 2.14,
        }
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_input_mode": "%",
                "ratio_percent": 20,
                "addition_sequence": 1,
            },
        ]

        workbook = self.generate_workbook(form_data=form_data, material_rows=material_rows)
        formula_sheet = workbook["Formula"]

        self.assertEqual(formula_sheet.cell(row=52, column=FORMULA_MATERIAL_COLUMNS["ratio"]).value, 0.2)
        self.assertEqual(
            formula_sheet.cell(row=52, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).value,
            "=G52*($C$18-$F$18)*$E$50",
        )

    def test_casing_krosok_percent_formula_uses_dynamic_blend_share(self):
        form_data = deepcopy(self.form_data)
        form_data["bypass_material_lookup"] = True
        form_data["phase_metadata"]["Casing Krosok"] = {
            "nav_code": "GEN-CS-00000",
            "description": "CS_K-CK-LCDT-04_R00",
            "blend_ratio": 0.52,
            "application": 0.0283,
        }
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": f"GEN-R-{index:05d}",
                "item_name": f"RAJANGAN MATERIAL {index}",
                "dosage_mg_stick": 1.0,
                "addition_sequence": index,
            }
            for index in range(1, 20)
        ]
        material_rows.extend([
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-K-00001",
                "item_name": "KROSOK MATERIAL",
                "dosage_input_mode": "%",
                "ratio_percent": 2.42,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-T-00001",
                "item_name": "TOP FLAVOR MATERIAL",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
        ])

        workbook = self.generate_workbook(form_data=form_data, material_rows=material_rows)
        formula_sheet = workbook["Formula"]
        krosok_row = SECTION_ROW_RANGES["Casing Krosok"][0]

        self.assertEqual(krosok_row, 53)
        self.assertEqual(
            formula_sheet.cell(row=krosok_row, column=FORMULA_MATERIAL_COLUMNS["dosage_mg_stick"]).value,
            "=G53*($C$18-$F$18)*$E$51*($E$50/($E$23+$E$50))",
        )
        self.assertEqual(formula_sheet.row_dimensions[krosok_row - 1].height, 48.0)

    def test_formula_output_is_sorted_by_addition_sequence_within_phase(self):
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 20,
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.25,
                "addition_sequence": 10,
            },
        ]
        workbook = self.generate_workbook(material_rows=self.with_required_main_phases(material_rows))
        formula_sheet = workbook["Formula"]

        self.assertEqual(formula_sheet.cell(row=26, column=2).value, "GEN-WA-00001")
        self.assertEqual(formula_sheet.cell(row=26, column=13).value, 10)
        self.assertEqual(formula_sheet.cell(row=27, column=2).value, "GEN-CHM-00015")
        self.assertEqual(formula_sheet.cell(row=27, column=13).value, 20)

    def test_process_instruction_columns_merge_by_addition_sequence(self):
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
                "temperature": "Ambient",
                "agitation_rate": "500 RPM",
                "mixing_duration": "5 min",
                "work_instruction_override": "Start mixing.",
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-CHM-00016",
                "item_name": "SV-PROPYLENE GLYCOL",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
                "temperature": "Ambient",
                "agitation_rate": "500 RPM",
                "mixing_duration": "15 min",
                "work_instruction_override": "Mix till all solids are dissolved.",
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
        ]
        workbook = self.generate_workbook(material_rows=material_rows)
        formula_sheet = workbook["Formula"]
        merged_ranges = {str(cell_range) for cell_range in formula_sheet.merged_cells.ranges}

        for column_name in ("addition_sequence", "temperature", "agitation_rate", "mixing_duration", "work_instruction"):
            column = FORMULA_MATERIAL_COLUMNS[column_name]
            column_letter = openpyxl.utils.get_column_letter(column)
            self.assertIn(f"{column_letter}27:{column_letter}28", merged_ranges)

        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["addition_sequence"]).value, 2)
        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["temperature"]).value, "Ambient")
        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["agitation_rate"]).value, "500 RPM")
        self.assertEqual(formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["mixing_duration"]).value, "15 min")
        self.assertEqual(
            formula_sheet.cell(row=27, column=FORMULA_MATERIAL_COLUMNS["work_instruction"]).value,
            "Mix till all solids are dissolved.",
        )

    def test_invalid_phase_dosage_and_stick_per_mc_are_rejected(self):
        cases = [
            ("unsupported phase", self.form_data, [{**self.material_rows[0], "phase": "Unknown Phase"}]),
            ("zero dosage", self.form_data, [{**self.material_rows[0], "dosage_mg_stick": 0}]),
            ("negative dosage", self.form_data, [{**self.material_rows[0], "dosage_mg_stick": -1}]),
            ("zero stick_per_mc", {**self.form_data, "stick_per_mc": 0}, self.material_rows),
            ("negative stick_per_mc", {**self.form_data, "stick_per_mc": -100}, self.material_rows),
        ]

        for name, form_data, material_rows in cases:
            with self.subTest(name=name):
                with TemporaryDirectory() as temp_dir:
                    output_path = Path(temp_dir) / "invalid.xlsx"
                    with self.assertRaises(ValueError):
                        formulation = self.build_formulation(form_data, material_rows)
                        generate_formulation_workbook(formulation, self.template_path, output_path)

    def test_missing_material_lookup_can_be_bypassed(self):
        material_rows = deepcopy(self.material_rows)
        material_rows[0]["item_code"] = "GEN-UNKNOWN-99999"
        material_rows[0]["item_name"] = "MATERIAL NOT IN MASTER"

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "missing_lookup.xlsx"
            with self.assertRaises(ValueError):
                generate_formulation_workbook(
                    self.build_formulation(material_rows=material_rows),
                    self.template_path,
                    output_path,
                )

        form_data = {**self.form_data, "bypass_material_lookup": True}
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "bypassed_lookup.xlsx"
            generate_formulation_workbook(
                self.build_formulation(form_data=form_data, material_rows=material_rows),
                self.template_path,
                output_path,
            )
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]

        first_casing_row = SECTION_ROW_RANGES["Casing Rajangan"][0]
        self.assertEqual(formula_sheet.cell(row=first_casing_row, column=2).value, "GEN-UNKNOWN-99999")
        self.assertEqual(formula_sheet.cell(row=first_casing_row, column=3).value, "MATERIAL NOT IN MASTER")

    def test_top_flavor_phase_expands_when_materials_exceed_base_capacity(self):
        top_flavor_rows = [
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-CHM-00016",
                "item_name": "SV-PROPYLENE GLYCOL",
                "dosage_mg_stick": 3.0,
                "addition_sequence": 3,
            },
        ]
        formulation = self.build_formulation(material_rows=self.with_required_main_phases(top_flavor_rows))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "top_flavor_expanded.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]
            top_flavor_rows_written = SECTION_ROW_RANGES["Top Flavor"][: len(top_flavor_rows)]

            self.assertGreaterEqual(len(SECTION_ROW_RANGES["Top Flavor"]), len(top_flavor_rows))
            self.assertEqual(
                [formula_sheet.cell(row=row, column=2).value for row in top_flavor_rows_written],
                ["GEN-WA-00001", "GEN-CHM-00015", "GEN-CHM-00016"],
            )
            self.assertEqual(formula_sheet.cell(row=13, column=3).value, build_formulation_code_formula())

    def test_material_name_merge_summary_and_approval_are_repaired_after_large_dynamic_insert(self):
        material_rows = []
        for phase, count in (("Casing Rajangan", 19), ("Casing Krosok", 19), ("Top Flavor", 4)):
            for index in range(count):
                material_rows.append({
                    "phase": phase,
                    "item_code": f"GEN-X-{phase[:2]}-{index:05d}",
                    "item_name": f"{phase} MATERIAL {index}",
                    "dosage_mg_stick": 1.0 + index / 10,
                    "addition_sequence": index + 1,
                })
        form_data = {**self.form_data, "bypass_material_lookup": True}
        formulation = self.build_formulation(
            form_data=form_data,
            material_rows=material_rows,
        )

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "large_dynamic_insert.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]

        merged_ranges = {str(cell_range) for cell_range in formula_sheet.merged_cells.ranges}
        for phase in ("Casing Rajangan", "Casing Krosok", "Top Flavor"):
            for row in SECTION_ROW_RANGES[phase]:
                self.assertIn(f"C{row}:D{row}", merged_ranges)
            first_material_row = SECTION_ROW_RANGES[phase][0]
            self.assertEqual(formula_sheet.cell(row=first_material_row, column=3).alignment.horizontal, "left")

        self.assertEqual(formula_sheet.cell(row=47, column=1).value, "Casing Krosok")
        self.assertEqual(formula_sheet.cell(row=48, column=5).alignment.horizontal, "left")
        self.assertEqual(formula_sheet.cell(row=48, column=1).value, "Casing NAV Item Code")
        self.assertEqual(formula_sheet.cell(row=49, column=1).value, "Casing NAV Item Description")
        self.assertEqual(formula_sheet.cell(row=74, column=1).value, "Top Flavor")
        self.assertEqual(formula_sheet.cell(row=75, column=1).value, "Top Flavor NAV Item Code")
        self.assertEqual(formula_sheet.cell(row=76, column=1).value, "Top Flavor Item Description")
        self.assertNotIn(
            formula_sheet.cell(row=SECTION_ROW_RANGES["Casing Rajangan"][0], column=2).font.color.rgb,
            ("FFFF0000", "00FF0000"),
        )

        summary_row = None
        for row in range(1, formula_sheet.max_row + 1):
            value = formula_sheet.cell(row=row, column=2).value
            if isinstance(value, str) and "Casing & Top Flavor" in value:
                summary_row = row
                break
        self.assertIsNotNone(summary_row)
        self.assertEqual(formula_sheet.cell(row=summary_row + 1, column=2).value, "=E22")
        self.assertEqual(formula_sheet.cell(row=summary_row + 2, column=2).value, "=E49")
        self.assertEqual(formula_sheet.cell(row=summary_row + 3, column=2).value, "=E76")

        approval_row = None
        for row in range(1, formula_sheet.max_row + 1):
            if formula_sheet.cell(row=row, column=1).value == "Prepared By":
                approval_row = row
                break
        self.assertIsNotNone(approval_row)
        self.assertIn(f"A{approval_row}:E{approval_row}", merged_ranges)
        self.assertIn(f"F{approval_row}:L{approval_row}", merged_ranges)
        self.assertIn(f"M{approval_row}:Q{approval_row}", merged_ranges)
        self.assertIn(f"A{approval_row + 1}:E{approval_row + 1}", merged_ranges)
        self.assertIn(f"F{approval_row + 1}:L{approval_row + 1}", merged_ranges)
        self.assertIn(f"M{approval_row + 1}:Q{approval_row + 1}", merged_ranges)
        self.assertIn(f"A{approval_row + 2}:E{approval_row + 2}", merged_ranges)
        self.assertIn(f"F{approval_row + 2}:L{approval_row + 2}", merged_ranges)
        self.assertIn(f"M{approval_row + 2}:Q{approval_row + 2}", merged_ranges)
        for row in (approval_row + 8, approval_row + 9):
            for column in range(13, 18):
                border = formula_sheet.cell(row=row, column=column).border
                self.assertIsNone(border.left.style if border.left else None)
                self.assertIsNone(border.right.style if border.right else None)
                self.assertIsNone(border.top.style if border.top else None)
                self.assertIsNone(border.bottom.style if border.bottom else None)

    def test_premix_phase_expands_when_materials_exceed_base_capacity(self):
        premix_rows = []
        for index in range(12):
            premix_rows.append({
                "phase": "Casing Pre-Mix",
                "item_code": "GEN-WA-00001" if index % 2 == 0 else "GEN-CHM-00015",
                "item_name": "WATER" if index % 2 == 0 else "INVERT SUGAR",
                "dosage_mg_stick": 1.0 + index,
                "addition_sequence": index + 1,
            })
        formulation = self.build_formulation(material_rows=self.with_required_main_phases(premix_rows))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "premix_over_capacity.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]
            premix_rows_written = SECTION_ROW_RANGES["Casing Pre-Mix"][: len(premix_rows)]
            self.assertGreaterEqual(len(SECTION_ROW_RANGES["Casing Pre-Mix"]), len(premix_rows))
            self.assertEqual(
                [formula_sheet.cell(row=row, column=2).value for row in premix_rows_written[:3]],
                ["GEN-WA-00001", "GEN-CHM-00015", "GEN-WA-00001"],
            )
            first_row = SECTION_ROW_RANGES["Casing Pre-Mix"][0]
            self.assertEqual(formula_sheet.cell(row=first_row - 4, column=1).value, "Casing Pre-Mix Formulation")
            self.assertEqual(formula_sheet.cell(row=first_row - 3, column=1).value, "Casing Pre-Mix Formulation NAV Item Code")
            self.assertEqual(formula_sheet.cell(row=first_row - 2, column=1).value, "Casing Pre-Mix Formulation NAV Item Description")
            total_row = SECTION_ROW_RANGES["Casing Pre-Mix"][-1] + 1
            self.assertEqual(formula_sheet.cell(row=total_row, column=1).value, "Total")
            self.assertIsNone(formula_sheet.cell(row=total_row, column=1).border.right.style)
            self.assertIsNone(formula_sheet.cell(row=total_row, column=2).border.left.style)
            self.assertEqual(formula_sheet.cell(row=total_row, column=7).border.left.style, "thin")
            self.assertEqual(formula_sheet.cell(row=total_row, column=17).border.right.style, "thin")

    def test_additional_flavor_premix_phase_creates_dynamic_section(self):
        premix_rows = [
            {
                "phase": "Flavor Pre-Mix 3",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Flavor Pre-Mix 3",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 2.0,
                "addition_sequence": 2,
            },
        ]
        formulation = self.build_formulation(material_rows=self.with_required_main_phases(premix_rows))

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "flavor_premix_3.xlsx"
            generate_formulation_workbook(formulation, self.template_path, output_path)
            workbook = openpyxl.load_workbook(output_path, data_only=False)
            formula_sheet = workbook["Formula"]

            self.assertIn("Flavor Pre-Mix 3", SECTION_ROW_RANGES)
            first_row = SECTION_ROW_RANGES["Flavor Pre-Mix 3"][0]
            self.assertEqual(formula_sheet.cell(row=first_row, column=2).value, "GEN-WA-00001")
            self.assertEqual(formula_sheet.cell(row=first_row + 1, column=2).value, "GEN-CHM-00015")

    def test_casing_premix_level_2_links_to_level_1_parent_dosage_and_price(self):
        form_data = deepcopy(self.form_data)
        form_data["bypass_material_lookup"] = True
        form_data["phase_metadata"]["Casing Pre-Mix"] = {
            "nav_code": "GEN-PM-00001",
            "description": "CPM_TEST",
            "blend_ratio": None,
            "application": None,
        }
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-PM-00001",
                "item_name": "CPM_TEST",
                "dosage_mg_stick": 5.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Pre-Mix",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_input_mode": "%",
                "ratio_percent": 20.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Pre-Mix",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_input_mode": "%",
                "ratio_percent": 80.0,
                "addition_sequence": 2,
            },
        ]

        workbook = self.generate_workbook(form_data=form_data, material_rows=material_rows)
        formula_sheet = workbook["Formula"]
        rajangan_rows = SECTION_ROW_RANGES["Casing Rajangan"]
        parent_row = next(row for row in rajangan_rows if formula_sheet.cell(row=row, column=2).value == "GEN-PM-00001")
        premix_first_row = SECTION_ROW_RANGES["Casing Pre-Mix"][0]
        premix_total_row = SECTION_ROW_RANGES["Casing Pre-Mix"][-1] + 1

        self.assertEqual(formula_sheet.cell(row=premix_first_row, column=7).value, 0.2)
        self.assertEqual(formula_sheet.cell(row=premix_first_row, column=8).value, f"=G{premix_first_row}*$H${parent_row}")
        self.assertEqual(formula_sheet.cell(row=parent_row, column=9).value, f"=J{premix_total_row}")
        self.assertTrue(fill_rgb(formula_sheet.cell(row=parent_row, column=2)).endswith(PARENT_BOM_LEVEL1_FILL_COLOR))
        self.assertTrue(fill_rgb(formula_sheet.cell(row=parent_row, column=8)).endswith(PARENT_BOM_LEVEL1_FILL_COLOR))
        self.assertTrue(formula_sheet.cell(row=parent_row, column=9).protection.locked)
        self.assertEqual(formula_sheet.cell(row=premix_first_row, column=2).border.left.style, "thin")
        self.assertEqual(formula_sheet.cell(row=premix_first_row, column=2).border.right.style, "thin")
        self.assertEqual(formula_sheet.cell(row=premix_first_row, column=2).border.top.style, "thin")
        self.assertEqual(formula_sheet.cell(row=premix_first_row, column=2).border.bottom.style, "thin")

    def test_flavor_premix_level_2_links_to_top_flavor_parent(self):
        form_data = deepcopy(self.form_data)
        form_data["bypass_material_lookup"] = True
        form_data["phase_metadata"]["Flavor Pre-Mix 3"] = {
            "nav_code": "GEN-PM-00003",
            "description": "FPM_TEST",
            "blend_ratio": None,
            "application": None,
        }
        material_rows = [
            {
                "phase": "Casing Rajangan",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Casing Krosok",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Top Flavor",
                "item_code": "GEN-PM-00003",
                "item_name": "FPM_TEST",
                "dosage_mg_stick": 2.5,
                "addition_sequence": 1,
            },
            {
                "phase": "Flavor Pre-Mix 3",
                "item_code": "GEN-WA-00001",
                "item_name": "WATER",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 1,
            },
            {
                "phase": "Flavor Pre-Mix 3",
                "item_code": "GEN-CHM-00015",
                "item_name": "INVERT SUGAR",
                "dosage_mg_stick": 1.0,
                "addition_sequence": 2,
            },
        ]

        workbook = self.generate_workbook(form_data=form_data, material_rows=material_rows)
        formula_sheet = workbook["Formula"]
        top_rows = SECTION_ROW_RANGES["Top Flavor"]
        parent_row = next(row for row in top_rows if formula_sheet.cell(row=row, column=2).value == "GEN-PM-00003")
        premix_first_row = SECTION_ROW_RANGES["Flavor Pre-Mix 3"][0]
        premix_total_row = SECTION_ROW_RANGES["Flavor Pre-Mix 3"][-1] + 1

        self.assertEqual(formula_sheet.cell(row=premix_first_row, column=8).value, f"=G{premix_first_row}*$H${parent_row}")
        self.assertEqual(formula_sheet.cell(row=parent_row, column=9).value, f"=J{premix_total_row}")
        self.assertTrue(fill_rgb(formula_sheet.cell(row=parent_row, column=2)).endswith(PARENT_BOM_LEVEL1_FILL_COLOR))
        self.assertTrue(fill_rgb(formula_sheet.cell(row=parent_row, column=8)).endswith(PARENT_BOM_LEVEL1_FILL_COLOR))
        self.assertTrue(formula_sheet.cell(row=parent_row, column=9).protection.locked)

    def test_set_cell_value_targets_merged_range_anchor(self):
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.merge_cells("A1:C1")

        set_cell_value(ws, 1, 2, "Merged header")

        self.assertEqual(ws["A1"].value, "Merged header")


if __name__ == "__main__":
    unittest.main()
